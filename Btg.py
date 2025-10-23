import fitz
import os
import re
import pandas as pd
import time
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from main import LoaderDialog


# ──────────────────────────────────────────────────────────────────────────────
# 🔹 Função para versão Streamlit (Web)
# ──────────────────────────────────────────────────────────────────────────────
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    """
    Função compatível com a Central de Bancos Web (Streamlit)
    ---------------------------------------------------------
    - files: lista de caminhos de PDFs enviados
    - output_dir: pasta onde salvar os resultados
    - progress_cb: callback de progresso (0–100)
    - log_cb: callback de logs de status
    ---------------------------------------------------------
    """
    log_cb("Iniciando processamento dos extratos do BTG Pactual...")

    total = len(files)
    dfs = []

    for i, pdf_path in enumerate(files, start=1):
        nome = os.path.basename(pdf_path)
        log_cb(f"📄 Processando arquivo {i}/{total}: {nome}")
        progress_cb(int((i - 1) / total * 60))

        try:
            df = extrair_lancamentos_pdf(pdf_path)
            if not df.empty:
                dfs.append(df)
                log_cb(f"✅ {len(df)} lançamentos extraídos de {nome}")
            else:
                log_cb(f"⚠️ Nenhum lançamento encontrado em {nome}")
        except Exception as e:
            log_cb(f"❌ Erro ao processar {nome}: {str(e)}")

        time.sleep(0.2)
        progress_cb(int((i / total) * 80))

    if dfs:
        df_final = pd.concat(dfs, ignore_index=True)
        excel_path = os.path.join(output_dir, "BTG_Resultados.xlsx")
        salvar_em_excel(df_final, excel_path)
        log_cb(f"💾 Planilha salva em: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento válido encontrado nos PDFs.")

    progress_cb(100)
    log_cb("Processamento concluído com sucesso! 🚀")


# ──────────────────────────────────────────────────────────────────────────────
# Extração dos lançamentos
# ──────────────────────────────────────────────────────────────────────────────
def extrair_lancamentos_pdf(caminho_pdf):
    doc = fitz.open(caminho_pdf)
    linhas_extraidas = []

    for pagina in doc:
        blocks = pagina.get_text("blocks")
        for b in blocks:
            for linha in b[4].split('\n'):
                linha_limpa = linha.strip()
                if linha_limpa:
                    linhas_extraidas.append(linha_limpa)

    padrao_data = re.compile(r"^\d{2}/\d{2}/\d{4}$")
    padrao_valor = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}")

    blocos = []
    i = 0
    while i < len(linhas_extraidas):
        linha = linhas_extraidas[i]

        if padrao_data.match(linha):
            data = linha
            i += 1
            descricao_linhas = []
            valor = None

            while i < len(linhas_extraidas):
                atual = linhas_extraidas[i].strip()

                if padrao_data.match(atual):
                    break

                if padrao_valor.fullmatch(atual):
                    valor_raw = atual
                    try:
                        valor = float(valor_raw.replace(".", "").replace(",", "."))
                    except ValueError:
                        valor = None
                    i += 1
                    break
                else:
                    descricao_linhas.append(atual)
                    i += 1

            descricao_final = " ".join(descricao_linhas).strip()

            # Ignorar blocos com "saldo"
            if "saldo" in descricao_final.lower():
                continue

            if data and descricao_final and valor is not None:
                blocos.append((data, descricao_final, valor))
        else:
            i += 1

    df = pd.DataFrame(blocos, columns=[
        "Data lançamento", "Descrição do lançamento", "Entradas / Saídas (R$)"
    ])

    df["Data lançamento"] = pd.to_datetime(df["Data lançamento"], format="%d/%m/%Y", errors='coerce')
    df = df.dropna(subset=["Data lançamento"]).reset_index(drop=True)
    df["Data lançamento"] = df["Data lançamento"].dt.strftime("%d/%m/%Y")
    return df


# ──────────────────────────────────────────────────────────────────────────────
# Salvamento formatado em Excel
# ──────────────────────────────────────────────────────────────────────────────
def salvar_em_excel(df, caminho_pdf):
    caminho_excel = Path(caminho_pdf).with_suffix('.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = "Lançamentos"

    bold = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")
    money_fmt = "#,##0.00"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.freeze_panes = 'A2'

    # Cabeçalhos
    for col_idx, coluna in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=coluna)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Dados
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, valor in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = border

            if col_idx == 3 and isinstance(valor, float):
                cell.number_format = money_fmt
                cell.font = Font(color="0000FF" if valor >= 0 else "FF0000")
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 18

    wb.save(caminho_excel)
    return caminho_excel


# ──────────────────────────────────────────────────────────────────────────────
# Fluxo padrão PyQt5 (Desktop)
# ──────────────────────────────────────────────────────────────────────────────
def processar_pdf_custom(janela):
    while True:
        arquivos, _ = QFileDialog.getOpenFileNames(
            janela,
            "Selecione um ou mais extratos PDF do BTG",
            "",
            "Arquivos PDF (*.pdf)"
        )

        if not arquivos:
            break

        for caminho_pdf in arquivos:
            dialog = LoaderDialog(janela, janela.light_theme)
            dialog.show()
            QApplication.processEvents()

            try:
                dialog.atualizar_progresso(10)
                df = extrair_lancamentos_pdf(caminho_pdf)
                dialog.atualizar_progresso(60)
                salvar_em_excel(df, caminho_pdf)
                dialog.atualizar_progresso(100)

            except Exception as e:
                QMessageBox.critical(
                    janela, "Erro", f"Erro ao processar:\n{caminho_pdf}\n\n{str(e)}"
                )

            dialog.accept()

        continuar = janela.mostrar_confirmacao(
            "Concluído",
            "Todos os arquivos foram processados.\n\nDeseja selecionar outros arquivos?"
        )
        if not continuar:
            break
