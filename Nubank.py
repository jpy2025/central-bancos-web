# ==========================================================
# Módulo: Nubank.py
# Compatível com:
#   - Streamlit (função processar_pdf_streamlit)
#   - PyQt5 (função processar_pdf_custom)
# ==========================================================

import re
import fitz
import os
import time
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from main import LoaderDialog


# ==========================================================
# 🔹 Função auxiliar para extrair lançamentos do PDF
# ==========================================================
def extrair_dados_pdf(caminho_pdf):
    texto_total = ''
    with fitz.open(caminho_pdf) as doc:
        for pagina in doc:
            texto_total += pagina.get_text()

    linhas = texto_total.splitlines()
    dados = []
    data_atual = ''
    buffer_movimentacao = []
    movimentacao_valida = False
    tipo_movimentacao = None

    meses = {
        "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
        "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
        "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12"
    }

    padroes_ignorar = [
        r'NU PAGAMENTOS', r'CNPJ', r'EXTRATO', r'PÁGINA', r'VALORES EM R\$',
        r'SALDO (FINAL|INICIAL|DO PERÍODO)', r'AGÊNCIA.*CONTA', r'MOVIMENTAÇÕES -'
    ]

    frases_chave = [
        "transferência enviada", "pix recebido", "pagamento de boleto",
        "compra aprovada", "cartão", "débito automático", "transferência recebida",
        "recarga", "resgate", "aplicação", "pagamento realizado"
    ]

    for linha in linhas:
        linha = linha.strip()

        if "total de entradas" in linha.lower():
            tipo_movimentacao = "entrada"
            continue
        elif "total de saídas" in linha.lower():
            tipo_movimentacao = "saida"
            continue

        if not linha or any(re.search(p, linha, re.IGNORECASE) for p in padroes_ignorar):
            continue

        match_data = re.match(r'^(\d{2}) (\w{3}) (\d{4})$', linha)
        if match_data:
            dia, mes_abrev, ano = match_data.groups()
            mes_num = meses.get(mes_abrev.upper())
            if mes_num:
                data_atual = f"{dia}/{mes_num}/{ano}"
            continue

        match_valor = re.match(r'^-?\d{1,3}(?:\.\d{3})*,\d{2}$', linha)
        if match_valor and data_atual and movimentacao_valida:
            valor_str = linha.replace('.', '').replace(',', '.')
            try:
                valor = float(valor_str)
                if tipo_movimentacao == "saida":
                    valor *= -1
                descricao = ' '.join(buffer_movimentacao).strip()
                if descricao:
                    dados.append({
                        'Data': data_atual,
                        'Movimentações': descricao,
                        'Valor': valor
                    })
                buffer_movimentacao = []
                movimentacao_valida = False
            except ValueError:
                continue
        else:
            if any(palavra in linha.lower() for palavra in frases_chave):
                buffer_movimentacao = [linha]
                movimentacao_valida = True
            elif movimentacao_valida:
                buffer_movimentacao.append(linha)

    return dados


# ==========================================================
# 🔹 Função para salvar o Excel formatado (usada em ambos os modos)
# ==========================================================
def salvar_em_excel(dados, caminho_pdf):
    df = pd.DataFrame(dados)
    df['Valor'] = df['Valor'].round(2)
    df['Data'] = pd.to_datetime(
        df['Data'], format='%d/%m/%Y', errors='coerce').dt.strftime('%d/%m/%Y')

    caminho_excel = Path(caminho_pdf).with_suffix('.xlsx')
    df.to_excel(caminho_excel, index=False)

    wb = load_workbook(caminho_excel)
    ws = wb.active

    ws.freeze_panes = 'A2'

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    col_valor = df.columns.get_loc('Valor') + 1
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.col_idx == col_valor:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                    cor = "0000FF" if cell.value > 0 else "FF0000"
                    cell.font = Font(color=cor)
                    cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    for col in ws.columns:
        max_length = max(len(str(cell.value))
                         if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    ref_final = f"{get_column_letter(ws.max_column)}{ws.max_row}"
    tabela = Table(displayName="TabelaExtrato", ref=f"A1:{ref_final}")
    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)

    wb.save(caminho_excel)
    return caminho_excel


# ==========================================================
# 💻 Modo Desktop (PyQt5)
# ==========================================================
def processar_pdf_custom(janela):
    while True:
        arquivos, _ = QFileDialog.getOpenFileNames(
            janela, "Selecione um ou mais PDFs do extrato Nubank", "", "Arquivos PDF (*.pdf)"
        )
        if not arquivos:
            break

        for caminho_pdf in arquivos:
            dialog = LoaderDialog(janela, janela.light_theme)
            dialog.show()
            QApplication.processEvents()

            try:
                dialog.atualizar_progresso(10)

                dados = extrair_dados_pdf(caminho_pdf)
                dialog.atualizar_progresso(60)

                if not dados:
                    QMessageBox.warning(
                        janela, "Aviso", f"Nenhum dado encontrado no PDF:\n{caminho_pdf}")
                    dialog.atualizar_progresso(100)
                    dialog.accept()
                    continue

                salvar_em_excel(dados, caminho_pdf)
                dialog.atualizar_progresso(90)

            except Exception as e:
                QMessageBox.critical(
                    janela, "Erro", f"Erro ao processar o arquivo:\n{e}")
            finally:
                dialog.atualizar_progresso(100)
                dialog.accept()

        continuar = janela.mostrar_confirmacao(
            "Concluído",
            "Todos os arquivos selecionados foram processados.\n\nDeseja selecionar novos arquivos?"
        )
        if not continuar:
            break


# ==========================================================
# 🌐 Modo Web (Streamlit)
# ==========================================================
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    log_cb("Iniciando processamento de arquivos Nubank...")

    total = len(files)
    registros = []

    for i, pdf_path in enumerate(files, start=1):
        log_cb(f"Lendo arquivo {i}/{total}: {os.path.basename(pdf_path)}")
        time.sleep(0.3)
        progress_cb(int((i / total) * 70))

        try:
            dados = extrair_dados_pdf(pdf_path)
            if not dados:
                log_cb(
                    f"⚠️ Nenhum lançamento encontrado em {os.path.basename(pdf_path)}")
                continue

            registros.extend(dados)

        except Exception as e:
            log_cb(f"❌ Erro ao processar {os.path.basename(pdf_path)}: {e}")

    if registros:
        df = pd.DataFrame(registros)
        os.makedirs(output_dir, exist_ok=True)
        excel_path = os.path.join(output_dir, "Nubank_Resultados.xlsx")
        df.to_excel(excel_path, index=False)
        log_cb(f"✅ Planilha gerada: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento encontrado nos PDFs enviados.")

    progress_cb(100)
    log_cb("Processamento concluído com sucesso! 🚀")
    return True
