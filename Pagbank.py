# ==========================================================
# Módulo: Pagbank.py
# Compatível com:
#   - Streamlit (função processar_pdf_streamlit)
#   - PyQt5 (função processar_pdf_custom)
# ==========================================================

import os
import re
import time
import fitz
import pandas as pd
from pathlib import Path
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from main import LoaderDialog


# ==========================================================
# 🔹 Extração de lançamentos do PDF
# ==========================================================
def extrair_lancamentos(pdf_path):
    doc = fitz.open(pdf_path)
    blocos = []

    for page in doc:
        blocos.extend(page.get_text("blocks"))

    dados = []
    for bloco in blocos:
        texto = bloco[4].strip()

        if not texto or "Saldo do dia" in texto or "Descrição" in texto:
            continue

        linhas = texto.split("\n")
        if len(linhas) < 2:
            continue

        if re.match(r"\d{2}/\d{2}/\d{4}", linhas[0]):
            data = linhas[0]
            valor = None

            for l in reversed(linhas):
                if "R$" in l:
                    valor = l.strip()
                    break

            if valor:
                descricao = " ".join(
                    l.strip() for l in linhas[1:] if l.strip() and l.strip() != valor
                )
                dados.append({
                    "Data": data,
                    "Descrição": descricao,
                    "Valor": valor
                })

    return pd.DataFrame(dados)


# ==========================================================
# 🔹 Formatação e salvamento do Excel
# ==========================================================
def salvar_em_excel_com_formatacao(df, pdf_path):
    def formatar_valor(valor_str):
        valor_str = valor_str.replace("R$", "").replace(
            " ", "").replace(".", "").replace(",", ".")
        return float(valor_str)

    df["Valor"] = df["Valor"].apply(formatar_valor)

    excel_path = os.path.splitext(pdf_path)[0] + ".xlsx"
    df.to_excel(excel_path, index=False)

    wb = load_workbook(excel_path)
    ws = wb.active

    # ✅ Congelar cabeçalho
    ws.freeze_panes = 'A2'

    # ✅ Estilo do cabeçalho
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ✅ Bordas finas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    col_valor = df.columns.get_loc('Valor') + 1

    # ✅ Formatação e cor do valor
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.col_idx == col_valor:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                    cor = "0000FF" if cell.value > 0 else "FF0000"
                    cell.font = Font(color=cor)
                    cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # ✅ Ajuste automático da largura das colunas
    for col in ws.columns:
        max_length = max(len(str(cell.value))
                         if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    # ✅ Criar tabela azul clara
    ref_final = f"{get_column_letter(ws.max_column)}{ws.max_row}"
    tabela = Table(displayName="TabelaPagbank", ref=f"A1:{ref_final}")
    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)

    wb.save(excel_path)
    return excel_path


# ==========================================================
# 💻 Modo Desktop (PyQt5)
# ==========================================================
def processar_pdf_custom(janela):
    while True:
        arquivos, _ = QFileDialog.getOpenFileNames(
            janela,
            "Selecione um ou mais extratos PagBank (PDF)",
            "",
            "Arquivos PDF (*.pdf)"
        )
        if not arquivos:
            break

        for file_path in arquivos:
            dialog = LoaderDialog(janela, janela.light_theme)
            dialog.show()
            QApplication.processEvents()

            try:
                dialog.atualizar_progresso(10)

                df = extrair_lancamentos(file_path)
                dialog.atualizar_progresso(50)

                if df.empty:
                    QMessageBox.warning(
                        janela, "Atenção", f"Nenhum lançamento encontrado no arquivo:\n{file_path}")
                    dialog.atualizar_progresso(100)
                    dialog.accept()
                    continue

                salvar_em_excel_com_formatacao(df, file_path)
                dialog.atualizar_progresso(90)

            except Exception as e:
                QMessageBox.critical(janela, "Erro", f"Ocorreu um erro:\n{e}")
            finally:
                dialog.atualizar_progresso(100)
                dialog.accept()

        continuar = janela.mostrar_confirmacao(
            "Concluído",
            "Todos os arquivos foram processados.\n\nDeseja selecionar novos arquivos?"
        )
        if not continuar:
            break


# ==========================================================
# 🌐 Modo Web (Streamlit)
# ==========================================================
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    log_cb("Iniciando processamento de arquivos PagBank...")

    total = len(files)
    registros = []

    for i, pdf_path in enumerate(files, start=1):
        log_cb(f"Lendo arquivo {i}/{total}: {os.path.basename(pdf_path)}")
        time.sleep(0.3)
        progress_cb(int((i / total) * 70))

        try:
            df = extrair_lancamentos(pdf_path)
            if df.empty:
                log_cb(
                    f"⚠️ Nenhum lançamento encontrado em {os.path.basename(pdf_path)}")
                continue

            registros.extend(df.to_dict('records'))

        except Exception as e:
            log_cb(f"❌ Erro ao processar {os.path.basename(pdf_path)}: {e}")

    if registros:
        df_final = pd.DataFrame(registros)
        os.makedirs(output_dir, exist_ok=True)
        excel_path = os.path.join(output_dir, "PagBank_Resultados.xlsx")
        df_final.to_excel(excel_path, index=False)
        log_cb(f"✅ Planilha gerada: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento encontrado nos PDFs enviados.")

    progress_cb(100)
    log_cb("Processamento concluído com sucesso! 🚀")
    return True
