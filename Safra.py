# ==========================================================
# Módulo: Safra.py
# Compatível com:
#   - Streamlit (função processar_pdf_streamlit)
#   - PyQt5 (função processar_pdf_custom)
# ==========================================================

import os
import re
import time
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from pathlib import Path
from main import LoaderDialog


# ==========================================================
# 🔹 Extração dos lançamentos do PDF Safra
# ==========================================================
def extrair_lancamentos_safra(caminho_pdf):
    lancamentos = []
    padrao_data = re.compile(r"^(\d{2}/\d{2})")
    padrao_valor = re.compile(r"^-?[\d\.]+,[\d]{2}$")
    ano_extrato = "2025"

    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            palavras = pagina.extract_words(use_text_flow=True)
            linha = []

            for palavra in palavras:
                if padrao_data.match(palavra['text']):
                    if linha:
                        linha_texto = " ".join(linha).lower()
                        if not ignorar_linha(linha_texto):
                            lancamentos += processar_linha(
                                linha, padrao_valor, ano_extrato)
                        linha = []
                linha.append(palavra['text'])

            if linha:
                linha_texto = " ".join(linha).lower()
                if not ignorar_linha(linha_texto):
                    lancamentos += processar_linha(
                        linha, padrao_valor, ano_extrato)

    return lancamentos


# ==========================================================
# 🔹 Função auxiliar para ignorar cabeçalhos e saldos
# ==========================================================
def ignorar_linha(linha_texto):
    cabecalhos = [
        "saldo + limite disponível", "saldo bloqueado", "limite cheque",
        "lançamentos realizados", "data lançamento complemento",
        "saldo", "conta corrente", "pagamento de bloqueto",
        "aplicacao cdb", "safra pay sa", "transf entre contas mesmo cpf"
    ]
    return "conta corrente" in linha_texto or any(c in linha_texto for c in cabecalhos)


# ==========================================================
# 🔹 Processamento individual de cada linha
# ==========================================================
def processar_linha(linha, padrao_valor, ano_extrato):
    resultados = []
    data = linha[0] + f"/{ano_extrato}"
    valor = None
    descricao = []
    ignorar = ["Banco", "Safra", "CNPJ", "Página", "Saldo"]

    for item in linha[1:]:
        if any(palavra.lower() in item.lower() for palavra in ignorar):
            continue
        if padrao_valor.match(item.replace(".", "")):
            valor = item.replace('.', '').replace(',', '.')
        else:
            descricao.append(item)

    if valor is not None:
        descricao_completa = " ".join(descricao)
        descricao_limpa = re.sub(r'\b\d+\s+de\s+\d+\b', '', descricao_completa)
        descricao_limpa = re.sub(
            r'S/A\s*58\.160\.789/0001-28', '', descricao_limpa)
        descricao_limpa = re.sub(r'\b\d{9}\b', '', descricao_limpa)
        descricao_limpa = re.sub(r'\s{2,}', ' ', descricao_limpa).strip()
        valor_float = float(valor.replace(
            '-', '')) if '-' not in valor else -float(valor.replace('-', ''))
        resultados.append([data, descricao_limpa, round(valor_float, 2)])

    return resultados


# ==========================================================
# 🔹 Salvamento e formatação Excel
# ==========================================================
def salvar_excel(dados, caminho_pdf):
    df = pd.DataFrame(dados, columns=["Data", "Descrição", "Valor (R$)"])
    caminho_final = Path(caminho_pdf).with_suffix('.xlsx')
    df.to_excel(caminho_final, index=False)

    wb = load_workbook(caminho_final)
    ws = wb.active

    # Congelar cabeçalho
    ws.freeze_panes = 'A2'

    # Estilo cabeçalho
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Bordas e cores
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            valor = cell.value
            cell.border = thin_border
            if isinstance(valor, (int, float)):
                cell.number_format = '#,##0.00'
                cell.font = Font(color="FF0000" if valor < 0 else "0000FF")
                cell.alignment = Alignment(horizontal="right")

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left")

    for col in ws.columns:
        max_len = max(len(str(cell.value))
                      if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(
            col[0].column)].width = max_len + 2

    ref_final = f"{get_column_letter(ws.max_column)}{ws.max_row}"
    tabela = Table(displayName="TabelaSafra", ref=f"A1:{ref_final}")
    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)
    wb.save(caminho_final)

    return caminho_final


# ==========================================================
# 💻 Modo Desktop (PyQt5)
# ==========================================================
def processar_pdf_custom(janela):
    while True:
        arquivos, _ = QFileDialog.getOpenFileNames(
            janela, "Selecione um ou mais extratos Safra (PDF)", "", "Arquivos PDF (*.pdf)"
        )
        if not arquivos:
            break

        for caminho_pdf in arquivos:
            dialog = LoaderDialog(janela, janela.light_theme)
            dialog.show()
            QApplication.processEvents()

            try:
                dialog.atualizar_progresso(10)
                dados = extrair_lancamentos_safra(caminho_pdf)
                dialog.atualizar_progresso(60)

                if not dados:
                    QMessageBox.warning(
                        janela, "Aviso", f"Nenhum lançamento encontrado em:\n{caminho_pdf}")
                    dialog.atualizar_progresso(100)
                    dialog.accept()
                    continue

                salvar_excel(dados, caminho_pdf)
                dialog.atualizar_progresso(100)

            except Exception as e:
                QMessageBox.critical(janela, "Erro", f"Ocorreu um erro:\n{e}")
                dialog.atualizar_progresso(100)

            dialog.accept()

        continuar = janela.mostrar_confirmacao(
            "Concluído",
            "Todos os arquivos foram processados.\n\nDeseja selecionar novos arquivos?"
        )
        if not continuar:
            break


# ==========================================================
# 🌐 Modo Web (Streamlit)
# ==========================================================
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    log_cb("Iniciando processamento de arquivos Safra...")

    total = len(files)
    registros = []

    for i, pdf_path in enumerate(files, start=1):
        log_cb(f"Lendo arquivo {i}/{total}: {os.path.basename(pdf_path)}")
        time.sleep(0.3)
        progress_cb(int((i / total) * 70))

        try:
            dados = extrair_lancamentos_safra(pdf_path)
            if not dados:
                log_cb(
                    f"⚠️ Nenhum lançamento encontrado em {os.path.basename(pdf_path)}")
                continue

            registros.extend(dados)

        except Exception as e:
            log_cb(f"❌ Erro ao processar {os.path.basename(pdf_path)}: {e}")

    if registros:
        df = pd.DataFrame(registros, columns=[
                          "Data", "Descrição", "Valor (R$)"])
        os.makedirs(output_dir, exist_ok=True)
        excel_path = os.path.join(output_dir, "Safra_Resultados.xlsx")
        df.to_excel(excel_path, index=False)
        log_cb(f"✅ Planilha gerada: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento encontrado nos PDFs enviados.")

    progress_cb(100)
    log_cb("Processamento concluído com sucesso! 🚀")
    return True
