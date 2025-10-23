import os
import re
import fitz
import pandas as pd
import time
from PyQt5.QtWidgets import QFileDialog, QApplication
from main import LoaderDialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo


# ══════════════════════════════════════════════════════════════════════════════
# 🔹 Função compatível com Central de Bancos Web (Streamlit)
# ══════════════════════════════════════════════════════════════════════════════
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    """
    Compatível com Central de Bancos Web (Streamlit)
    ---------------------------------------------------------
    - files: lista de PDFs enviados
    - output_dir: pasta onde salvar o Excel
    - progress_cb: callback (0–100)
    - log_cb: função de log
    ---------------------------------------------------------
    """
    log_cb("Iniciando processamento dos extratos do Itaú Manix...")
    total = len(files)
    all_dataframes = []

    for i, caminho_pdf in enumerate(files, start=1):
        nome = os.path.basename(caminho_pdf)
        log_cb(f"📄 Lendo arquivo {i}/{total}: {nome}")
        progress_cb(int((i - 1) / total * 60))

        try:
            df = extrair_lancamentos_itau(caminho_pdf)
            if not df.empty:
                all_dataframes.append(df)
                log_cb(f"✅ {len(df)} lançamentos extraídos de {nome}")
            else:
                log_cb(f"⚠️ Nenhum lançamento encontrado em {nome}")
        except Exception as e:
            log_cb(f"❌ Erro ao processar {nome}: {e}")

        time.sleep(0.2)
        progress_cb(int((i / total) * 80))

    if all_dataframes:
        df_final = pd.concat(all_dataframes, ignore_index=True)
        excel_path = os.path.join(output_dir, "Itau_Manix_Resultados.xlsx")
        df_final.to_excel(excel_path, index=False)
        aplicar_formatacao_excel(excel_path)
        log_cb(f"💾 Planilha salva em: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento válido encontrado.")

    progress_cb(100)
    log_cb("✅ Processamento concluído com sucesso! 🚀")


# ══════════════════════════════════════════════════════════════════════════════
# Extração dos lançamentos do Itaú Manix (versão condensada e robusta)
# ══════════════════════════════════════════════════════════════════════════════
def extrair_lancamentos_itau(caminho_pdf):
    doc = fitz.open(caminho_pdf)
    linhas = []

    # Extração e filtragem de linhas
    for i, page in enumerate(doc):
        texto = page.get_text("text")
        linhas_pagina = [l for l in texto.split(
            '\n') if not re.match(r'^ {2,}', l)]

        if i == len(doc) - 1:
            padrao_saldo = re.compile(r'^SALDO \d{2}/\d{2}/\d{4}$')
            for j in range(len(linhas_pagina) - 1, -1, -1):
                if padrao_saldo.match(linhas_pagina[j]):
                    linhas_pagina = linhas_pagina[:j]
                    break
        linhas.extend(linhas_pagina)

    doc.close()

    # Limpa ruídos
    linhas = [l.strip() for l in linhas if l.strip() and not (
        any(re.search(padrao, l, re.IGNORECASE) for padrao in [
            r"^MANIX", r"^Data:", r"^Hora:", r"^Pág\.", r"Financeiro Extrato",
            r"^Período:", r"^SALDO", r"TOTAL", r"EXTRATO DE CONTA"
        ])
    )]

    # Identifica o mês predominante
    datas_detectadas = [re.match(r"\d{2}/\d{2}/\d{4}", l)
                        for l in linhas if re.match(r"\d{2}/\d{2}/\d{4}", l)]
    datas_convertidas = [pd.to_datetime(m.group(), dayfirst=True)
                         for m in datas_detectadas if m]
    if not datas_convertidas:
        return pd.DataFrame()

    contagem = pd.Series((d.month, d.year)
                         for d in datas_convertidas).value_counts()
    mes_referencia, ano_referencia = contagem.idxmax()

    # Monta os lançamentos
    lancamentos = []
    i = 0
    while i < len(linhas):
        linha = linhas[i]
        if re.match(r"\d{2}/\d{2}/\d{4}", linha):
            emissao = linha.strip()
            origem = conta = observacao = cheque = valor = ""
            bloco = [emissao]
            i += 1
            while i < len(linhas) and not re.match(r"\d{2}/\d{2}/\d{4}", linhas[i]):
                bloco.append(linhas[i])
                i += 1

            bloco_unido = " ".join(bloco).lower()
            if any(p in bloco_unido for p in ["saldo", "total crédito", "total débito"]):
                continue

            for b in bloco[1:]:
                b_limpo = b.strip()
                if re.match(r"^(LANC-|TESOUR.|LOT)", b) and not origem:
                    origem = b_limpo
                elif re.match(r"-?\d{1,3}(\.\d{3})*,\d{2}$", b) and not valor:
                    valor = b_limpo
                elif re.match(r"^\d{5,8}$", b) and not cheque:
                    cheque = b_limpo
                elif not conta:
                    conta = b_limpo
                else:
                    conta += " " + b_limpo

            try:
                data_dt = pd.to_datetime(emissao, dayfirst=True)
                if (data_dt.month != mes_referencia or data_dt.year != ano_referencia):
                    continue
            except:
                continue

            if emissao and valor:
                valor_limpo = valor.replace(".", "").replace(",", ".")
                lancamentos.append({
                    "EMISSÃO": emissao,
                    "ORIGEM": origem,
                    "CONTA/FORMA PGTO": conta.strip(),
                    "OBSERVAÇÃO": observacao.strip(),
                    "CHEQUE": cheque,
                    "VALOR": float(valor_limpo)
                })
        else:
            i += 1

    return pd.DataFrame(lancamentos)


# ══════════════════════════════════════════════════════════════════════════════
# Formatação do Excel (estilo azul e condicional)
# ══════════════════════════════════════════════════════════════════════════════
def aplicar_formatacao_excel(caminho_excel):
    wb = load_workbook(caminho_excel)
    ws = wb.active

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color="1F4E78",
                              end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border

    # Ajuste largura
    for col in ws.columns:
        max_len = max(len(str(cell.value))
                      if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    # Coluna de valor
    col_valor_index = [cell.value for cell in ws[1]].index("VALOR") + 1
    valor_col_letter = ws.cell(row=1, column=col_valor_index).column_letter

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=col_valor_index, max_col=col_valor_index):
        for cell in row:
            cell.number_format = '#,##0.00'

    # Condicional: azul e vermelho
    ws.conditional_formatting.add(
        f"{valor_col_letter}2:{valor_col_letter}{ws.max_row}",
        CellIsRule(operator='lessThan', formula=[
                   '0'], stopIfTrue=True, font=Font(color="FF0000"))
    )
    ws.conditional_formatting.add(
        f"{valor_col_letter}2:{valor_col_letter}{ws.max_row}",
        CellIsRule(operator='greaterThanOrEqual', formula=[
                   '0'], stopIfTrue=True, font=Font(color="0000FF"))
    )

    # Tabela estilizada
    table_ref = f"A1:{valor_col_letter}{ws.max_row}"
    table = Table(displayName="TabelaDados", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    ws.freeze_panes = "A2"
    wb.save(caminho_excel)


# ══════════════════════════════════════════════════════════════════════════════
# Fluxo PyQt5 (Desktop)
# ══════════════════════════════════════════════════════════════════════════════
def processar_pdf_custom(janela):
    try:
        while True:
            caminhos_pdf, _ = QFileDialog.getOpenFileNames(
                janela, "Selecionar extratos Itaú Manix (PDF)", "", "Arquivos PDF (*.pdf)"
            )

            if not caminhos_pdf:
                return

            for caminho_pdf in caminhos_pdf:
                dialog = LoaderDialog(janela, janela.light_theme)
                dialog.show()
                QApplication.processEvents()
                dialog.atualizar_progresso(10)

                df = extrair_lancamentos_itau(caminho_pdf)
                dialog.atualizar_progresso(70)

                if df.empty:
                    dialog.accept()
                    janela.mostrar_mensagem(
                        "Atenção", f"Nenhum lançamento válido encontrado no arquivo:\n{os.path.basename(caminho_pdf)}"
                    )
                    continue

                nome_base = os.path.splitext(os.path.basename(caminho_pdf))[0]
                caminho_excel = os.path.join(
                    os.path.dirname(caminho_pdf), f"{nome_base}.xlsx")
                df.to_excel(caminho_excel, index=False)
                aplicar_formatacao_excel(caminho_excel)

                dialog.atualizar_progresso(100)
                dialog.accept()

            continuar = janela.mostrar_confirmacao(
                "Concluído", "Todos os arquivos selecionados foram processados.\n\nDeseja selecionar novos arquivos?"
            )
            if not continuar:
                break

    except Exception as e:
        janela.mostrar_mensagem(
            "Erro", f"Ocorreu um erro ao processar o PDF:\n{str(e)}")
