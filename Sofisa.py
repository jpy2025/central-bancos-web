# ==========================================================
# Módulo: Sofisa.py
# Compatível com:
#   - Streamlit (função processar_pdf_streamlit)
#   - PyQt5 (função processar_pdf_custom)
# ==========================================================

import os
import re
import time
import fitz  # PyMuPDF
import pandas as pd
from pdf2image import convert_from_path
import pytesseract
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from main import LoaderDialog


# ==========================================================
# 🔹 Extração dos lançamentos
# ==========================================================
def extrair_lancamentos(texto):
    linhas = texto.splitlines()
    dados = []
    grupo_atual = []

    def is_linha_invalida(linha):
        return (
            re.match(r"^Saldo disponível em \d{2}/\d{2}/\d{2}", linha)
            or "Atualizado em" in linha
            or "Extrato por período" in linha
            or "Entradas/Saídas" in linha
            or "Saldo em conta" in linha
            or "Valor Bloqueado" in linha
            or "Cheque Fácil" in linha
            or "Saldo Bloqueado" in linha
            or "Cliente :" in linha
            or "Agência:" in linha
            or "Conta:" in linha
            or re.search(r"https?://", linha)
            or re.match(r"^\d{2}/\d{2}/\d{4}$", linha)
            or re.match(r"^\d{2}/\d{2}/\d{4},\s*\d{2}:\d{2}", linha)
        )

    def extrair_valor(grupo):
        for linha in grupo:
            match = re.search(r"([-=]?)\s*(\d{1,3}(?:\.\d{3})*,\d{2})", linha)
            if match:
                sinal, valor_txt = match.groups()
                valor = float(valor_txt.replace('.', '').replace(',', '.'))
                return -abs(valor) if '-' in sinal or '=' in sinal else valor
        return None

    def remover_valores(grupo):
        return [
            re.sub(r"([-=]?)\s*\d{1,3}(?:\.\d{3})*,\d{2}", "", linha).strip()
            for linha in grupo
        ]

    def processar_grupo(grupo):
        if not grupo:
            return None
        match = re.match(r"^(\d{2}/\d{2}/\d{2})(.*)", grupo[0])
        if not match:
            return None
        data = pd.to_datetime(match.group(1), format='%d/%m/%y')
        grupo[0] = match.group(2).strip()
        valor = extrair_valor(grupo)
        grupo_limpo = remover_valores(grupo)
        descricao = " ".join(l for l in grupo_limpo if l).strip()
        descricao = re.sub(r"\s+", " ", descricao)
        return {
            "Data": data,
            "Lançamentos": descricao,
            "Valor (R$)": valor
        }

    i = 0
    while i < len(linhas):
        linha = linhas[i].strip()

        if not linha or is_linha_invalida(linha):
            i += 1
            continue

        if re.match(r"^\d{2}/\d{2}/\d{2}", linha):
            grupo_atual = [linha]
            i += 1
            while i < len(linhas):
                prox = linhas[i].strip()
                if is_linha_invalida(prox) or re.match(r"^\d{2}/\d{2}/\d{2}", prox):
                    break
                grupo_atual.append(prox)
                i += 1
            registro = processar_grupo(grupo_atual)
            if registro:
                dados.append(registro)
        else:
            i += 1

    return dados


# ==========================================================
# 🔹 Extração de texto (PDF direto ou OCR)
# ==========================================================
def extrair_texto_pdf_ou_ocr(caminho_pdf):
    texto_completo = ''
    try:
        doc = fitz.open(caminho_pdf)
        for pagina in doc:
            texto_completo += pagina.get_text()
        doc.close()
        if texto_completo.strip():
            return texto_completo
    except Exception as e:
        print("Erro ao usar PyMuPDF:", e)

    imagens = convert_from_path(caminho_pdf)
    for imagem in imagens:
        texto_completo += pytesseract.image_to_string(imagem, lang='por')
    return texto_completo


# ==========================================================
# 🔹 Formatação e salvamento Excel
# ==========================================================
def formatar_excel(caminho_excel):
    wb = load_workbook(caminho_excel)
    ws = wb.active

    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    col_valor = 3

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.col_idx == 1:
                cell.number_format = 'DD/MM/YYYY'
                cell.alignment = Alignment(horizontal="center")
            elif cell.col_idx == col_valor:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                    cor = "0000FF" if cell.value >= 0 else "FF0000"
                    cell.font = Font(color=cor)
                    cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value))
                         if cell.value else 0 for cell in column_cells)
        column_letter = column_cells[0].column_letter
        ws.column_dimensions[column_letter].width = max_length + 2

    num_linhas = ws.max_row
    ref_final = f"{get_column_letter(ws.max_column)}{num_linhas}"
    tab = Table(displayName="TabelaSofisa", ref=f"A1:{ref_final}")

    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = estilo
    ws.add_table(tab)

    wb.save(caminho_excel)
    return caminho_excel


# ==========================================================
# 💻 Modo Desktop (PyQt5)
# ==========================================================
def processar_pdf_custom(janela):
    while True:
        arquivos, _ = QFileDialog.getOpenFileNames(
            janela,
            "Selecione os PDFs do extrato Sofisa",
            "",
            "Arquivos PDF (*.pdf)"
        )
        if not arquivos:
            break

        for caminho_pdf in arquivos:
            dialog = LoaderDialog(janela, janela.light_theme)
            dialog.show()
            QApplication.processEvents()

            try:
                dialog.atualizar_progresso(10)

                texto = extrair_texto_pdf_ou_ocr(caminho_pdf)
                dialog.atualizar_progresso(40)

                dados = extrair_lancamentos(texto)
                dialog.atualizar_progresso(60)

                if not dados:
                    QMessageBox.warning(
                        janela, "Atenção", f"Nenhum dado foi extraído do PDF:\n{caminho_pdf}")
                    dialog.atualizar_progresso(100)
                    dialog.accept()
                    continue

                df = pd.DataFrame(dados)
                caminho_excel = os.path.splitext(caminho_pdf)[0] + '.xlsx'
                df.to_excel(caminho_excel, index=False)
                dialog.atualizar_progresso(80)

                formatar_excel(caminho_excel)
                dialog.atualizar_progresso(100)

            except Exception as e:
                QMessageBox.critical(
                    janela, "Erro", f"Ocorreu um erro ao processar {caminho_pdf}:\n{e}")
                dialog.atualizar_progresso(100)

            dialog.accept()

        continuar = janela.mostrar_confirmacao(
            "Concluído",
            "Todos os arquivos selecionados foram processados.\n\nDeseja selecionar novos arquivos?"
        )
        if not continuar:
            break


# ==========================================================
# 🌐 Modo Web (Streamlit)
# ==========================================================
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    log_cb("Iniciando processamento de extratos Sofisa...")

    total = len(files)
    registros = []

    for i, pdf_path in enumerate(files, start=1):
        log_cb(f"Lendo arquivo {i}/{total}: {os.path.basename(pdf_path)}")
        time.sleep(0.3)
        progress_cb(int((i / total) * 60))

        try:
            texto = extrair_texto_pdf_ou_ocr(pdf_path)
            dados = extrair_lancamentos(texto)
            if not dados:
                log_cb(
                    f"⚠️ Nenhum lançamento encontrado em {os.path.basename(pdf_path)}")
                continue

            registros.extend(dados)

        except Exception as e:
            log_cb(f"❌ Erro ao processar {os.path.basename(pdf_path)}: {e}")

    if registros:
        df_final = pd.DataFrame(registros)
        os.makedirs(output_dir, exist_ok=True)
        excel_path = os.path.join(output_dir, "Sofisa_Resultados.xlsx")
        df_final.to_excel(excel_path, index=False)
        log_cb(f"✅ Planilha gerada: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento encontrado nos PDFs enviados.")

    progress_cb(100)
    log_cb("Processamento concluído com sucesso! 🚀")
    return True
