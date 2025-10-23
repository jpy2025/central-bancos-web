import fitz
import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication

from main import LoaderDialog


# ==========================================================
# 🔹 Função de Extração
# ==========================================================
def extrair_lancamentos(texto):
    match_data = re.search(r"Mês:\s+([A-Za-zçÇ]+)[/\s](\d{4})", texto)
    if not match_data:
        raise ValueError("Mês e ano não encontrados.")
    mes_nome, ano = match_data.groups()
    meses = {
        'janeiro': '01', 'fevereiro': '02', 'março': '03', 'abril': '04',
        'maio': '05', 'junho': '06', 'julho': '07', 'agosto': '08',
        'setembro': '09', 'outubro': '10', 'novembro': '11', 'dezembro': '12'
    }
    mes = meses.get(mes_nome.lower())
    if not mes:
        raise ValueError(f"Mês inválido: {mes_nome}")

    linhas = texto.splitlines()
    lancamentos = []
    dia_atual = None

    padrao = re.compile(
        r"^\s*(?:(\d{1,2})\s+)?([A-ZÇÃÂÉÈÓÔÕÍ\s0-9\-\.]+?)\s+(\d{4,5})\s+([\d\.]+,\d{2})(\+|\-)"
    )

    for linha in linhas:
        match = padrao.search(linha)
        if match:
            dia, historico, documento, valor, sinal = match.groups()
            if dia:
                dia_atual = int(dia)
            if dia_atual is None:
                continue
            valor_float = float(valor.replace('.', '').replace(',', '.'))
            if sinal == "-":
                valor_float *= -1
            data = f"{dia_atual:02d}/{mes}/{ano}"
            lancamentos.append([data, historico.strip().title(), valor_float])

    return pd.DataFrame(lancamentos, columns=["Data", "Histórico", "Valor"])


# ==========================================================
# 🔹 Função de Salvamento Excel
# ==========================================================
def salvar_em_excel(df, pdf_path):
    pasta, nome = os.path.split(pdf_path)
    nome_base = os.path.splitext(nome)[0]
    caminho_excel = os.path.join(pasta, f"{nome_base}.xlsx")

    df.to_excel(caminho_excel, index=False)

    wb = load_workbook(caminho_excel)
    ws = wb.active

    ws.freeze_panes = 'A2'

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border
            if cell.column == 3:
                valor = cell.value
                if isinstance(valor, (int, float)):
                    cell.number_format = '#,##0.00'
                    if valor < 0:
                        cell.font = Font(color='FF0000')
                    else:
                        cell.font = Font(color='0000FF')
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="left")

    for col in ws.columns:
        max_len = max(len(str(cell.value))
                      if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 2

    num_linhas = ws.max_row
    tab = Table(displayName="TabelaLancamentos", ref=f"A1:C{num_linhas}")

    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = estilo
    ws.add_table(tab)

    wb.save(caminho_excel)
    return caminho_excel


# ==========================================================
# 🔹 Função Desktop (Central de Bancos original)
# ==========================================================
def processar_pdf_custom(qt_parent):
    try:
        while True:
            caminhos_pdf, _ = QFileDialog.getOpenFileNames(
                qt_parent, "Selecione um ou mais extratos BNB em PDF", "", "PDF Files (*.pdf)"
            )

            if not caminhos_pdf:
                break

            for caminho_pdf in caminhos_pdf:
                dialog = LoaderDialog(qt_parent, qt_parent.light_theme)
                dialog.show()
                QApplication.processEvents()

                try:
                    dialog.atualizar_progresso(10)

                    texto = ""
                    with fitz.open(caminho_pdf) as doc:
                        dialog.atualizar_progresso(30)
                        for page in doc:
                            texto += page.get_text()

                    dialog.atualizar_progresso(50)

                    df = extrair_lancamentos(texto)
                    dialog.atualizar_progresso(70)

                    if df.empty:
                        QMessageBox.warning(
                            qt_parent, "Aviso", f"Nenhum lançamento encontrado no arquivo:\n{os.path.basename(caminho_pdf)}")
                    else:
                        salvar_em_excel(df, caminho_pdf)
                        dialog.atualizar_progresso(90)

                except Exception as e:
                    QMessageBox.critical(
                        qt_parent, "Erro", f"Erro no arquivo {os.path.basename(caminho_pdf)}:\n{str(e)}")
                    dialog.atualizar_progresso(100)

                dialog.accept()

            continuar = qt_parent.mostrar_confirmacao(
                "Concluído",
                "Todos os arquivos selecionados foram processados.\n\nDeseja selecionar novos arquivos?"
            )
            if not continuar:
                break

    except Exception as erro_final:
        QMessageBox.critical(qt_parent, "Erro geral",
                             f"Erro inesperado:\n{str(erro_final)}")


# ==========================================================
# 🔹 Função Streamlit (Web-friendly)
# ==========================================================
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    """
    Versão web-friendly (usada no app Streamlit)
    """
    log_cb("Iniciando processamento de arquivos BNB...")
    total = len(files)
    registros = []

    for i, pdf_path in enumerate(files, start=1):
        log_cb(f"Lendo arquivo {i}/{total}: {os.path.basename(pdf_path)}")
        progress_cb(int((i / total) * 70))

        try:
            with fitz.open(pdf_path) as doc:
                texto = "\n".join(page.get_text() for page in doc)
            df = extrair_lancamentos(texto)
            if not df.empty:
                registros.append(df)
        except Exception as e:
            log_cb(f"❌ Erro ao processar {os.path.basename(pdf_path)}: {e}")

    if registros:
        df_final = pd.concat(registros, ignore_index=True)
        os.makedirs(output_dir, exist_ok=True)
        excel_path = os.path.join(output_dir, "BNB_Resultados.xlsx")
        df_final.to_excel(excel_path, index=False)
        log_cb(f"✅ Planilha gerada: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento encontrado.")

    progress_cb(100)
    log_cb("Processamento concluído com sucesso! 🚀")
    return True
