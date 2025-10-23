import re
import os
import time
import fitz  # PyMuPDF
import pandas as pd
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows


# ✅ Importa LoaderDialog
from main import LoaderDialog


# ──────────────────────────────────────────────────────────────────────────────
# 🔹 Função compatível com o modo Web (Streamlit)
# ──────────────────────────────────────────────────────────────────────────────
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    """
    Função compatível com a Central de Bancos Web (Streamlit)
    ---------------------------------------------------------
    - files: lista de PDFs enviados
    - output_dir: pasta onde salvar o Excel
    - progress_cb: callback de progresso (0–100)
    - log_cb: função de log de mensagens
    ---------------------------------------------------------
    """
    log_cb("Iniciando processamento dos extratos do Banco Daycoval...")

    total = len(files)
    todos_dados = []

    for i, pdf_path in enumerate(files, start=1):
        nome = os.path.basename(pdf_path)
        log_cb(f"📄 Lendo arquivo {i}/{total}: {nome}")
        progress_cb(int((i - 1) / total * 60))

        try:
            dados = extrair_lancamentos(pdf_path)
            if dados:
                todos_dados.extend(dados)
                log_cb(f"✅ {len(dados)} lançamentos extraídos de {nome}")
            else:
                log_cb(f"⚠️ Nenhum lançamento encontrado em {nome}")
        except Exception as e:
            log_cb(f"❌ Erro ao processar {nome}: {str(e)}")

        time.sleep(0.2)
        progress_cb(int((i / total) * 80))

    if todos_dados:
        df = pd.DataFrame(todos_dados)
        df = df[~df["Lançamento"].str.lower().str.contains("saldo")]

        excel_path = os.path.join(output_dir, "Daycoval_Resultados.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Lançamentos"
        ws.append(["Data", "Lançamento", "Valor"])

        for _, row in df.iterrows():
            data = row["Data"]
            desc = row["Lançamento"]
            valor_str = row["Valor"]
            valor_num = float(valor_str.replace(".", "").replace(",", "."))
            ws.append([data, desc, valor_num])

        _formatar_excel(ws)
        wb.save(excel_path)
        log_cb(f"💾 Planilha salva em: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento válido encontrado nos PDFs.")

    progress_cb(100)
    log_cb("✅ Processamento concluído com sucesso! 🚀")


# ──────────────────────────────────────────────────────────────────────────────
# Extração dos lançamentos
# ──────────────────────────────────────────────────────────────────────────────
def extrair_lancamentos(pdf_path):
    doc = fitz.open(pdf_path)
    lancamentos = []
    ano_extrato = None
    data_atual = None
    buffer_lancamento = []

    padrao_valor = re.compile(r'-?\s?\d{1,3}(?:\.\d{3})*,\d{2}$')
    padrao_data = re.compile(r'^\d{2}/\d{2}$')
    padrao_data_completa = re.compile(r'\d{2}/\d{2}/(\d{4})')
    padrao_linha_unica = re.compile(
        r'^(\d{2}/\d{2})\s+(.*?)\s+(-?\s?\d{1,3}(?:\.\d{3})*,\d{2})$')

    linhas_todas = []
    for page in doc:
        linhas = page.get_text("text").split('\n')
        if len(linhas) > 6:
            linhas = linhas[3:-3]
        linhas_todas.extend([linha.strip()
                            for linha in linhas if linha.strip()])
    doc.close()

    for linha in linhas_todas:
        if not ano_extrato:
            match_ano = padrao_data_completa.search(linha)
            if match_ano:
                ano_extrato = match_ano.group(1)

        match_linha_unica = padrao_linha_unica.match(linha)
        if match_linha_unica:
            data_atual, descricao, valor = match_linha_unica.groups()
            valor = valor.replace(" ", "")
            if not descricao or valor == "0,00":
                continue
            data_formatada = f"{data_atual}/{ano_extrato}" if ano_extrato else data_atual
            lancamentos.append({
                "Data": data_formatada,
                "Lançamento": descricao.strip(),
                "Valor": valor
            })
            continue

        if padrao_data.fullmatch(linha):
            data_atual = linha
            buffer_lancamento = []
            continue

        if padrao_valor.fullmatch(linha):
            if data_atual:
                valor = linha.replace(" ", "")
                descricao = ' '.join(buffer_lancamento).strip()
                if not descricao or valor == "0,00":
                    buffer_lancamento = []
                    continue
                data_formatada = f"{data_atual}/{ano_extrato}" if ano_extrato else data_atual
                lancamentos.append({
                    "Data": data_formatada,
                    "Lançamento": descricao,
                    "Valor": valor
                })
            buffer_lancamento = []
            continue

        buffer_lancamento.append(linha)

    return lancamentos


# ──────────────────────────────────────────────────────────────────────────────
# Formatação padrão do Excel
# ──────────────────────────────────────────────────────────────────────────────
def _formatar_excel(ws):
    # Cabeçalho
    ws.freeze_panes = 'A2'
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Dados
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border
            if cell.col_idx == 3:
                cell.number_format = '#,##0.00'
                cell.font = Font(
                    color="0000FF" if cell.value >= 0 else "FF0000")
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Ajuste de largura
    for col in ws.columns:
        max_len = max(len(str(cell.value))
                      if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 2

    # Tabela estilizada
    num_linhas = ws.max_row
    tab = Table(displayName="TabelaLancamentos", ref=f"A1:C{num_linhas}")
    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = estilo
    ws.add_table(tab)


# ──────────────────────────────────────────────────────────────────────────────
# Fluxo padrão PyQt5 (Desktop)
# ──────────────────────────────────────────────────────────────────────────────
def processar_pdf_custom(qt_parent):
    while True:
        arquivos, _ = QFileDialog.getOpenFileNames(
            qt_parent, "Selecione um ou mais extratos do Daycoval", "", "PDF Files (*.pdf)"
        )
        if not arquivos:
            break

        for file_path in arquivos:
            dialog = LoaderDialog(qt_parent, qt_parent.light_theme)
            dialog.show()
            QApplication.processEvents()

            try:
                dialog.atualizar_progresso(10)
                dados = extrair_lancamentos(file_path)
                dialog.atualizar_progresso(60)

                if dados:
                    df = pd.DataFrame(dados)
                    df = df[~df["Lançamento"].str.lower().str.contains("saldo")]
                    excel_path = os.path.splitext(file_path)[0] + ".xlsx"
                    df.to_excel(excel_path, index=False)
                    dialog.atualizar_progresso(80)

                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Lançamentos"
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)

                    _formatar_excel(ws)
                    wb.save(excel_path)
                    dialog.atualizar_progresso(100)

                else:
                    QMessageBox.warning(
                        qt_parent, "Aviso",
                        f"Nenhum lançamento encontrado no arquivo:\n{os.path.basename(file_path)}"
                    )
                    dialog.atualizar_progresso(100)

            except Exception as e:
                QMessageBox.critical(
                    qt_parent, "Erro",
                    f"Ocorreu um erro no arquivo {os.path.basename(file_path)}:\n{e}"
                )
                dialog.atualizar_progresso(100)

            dialog.accept()

            continuar = qt_parent.mostrar_confirmacao(
                "Concluído",
                "Todos os arquivos selecionados foram processados.\n\nDeseja selecionar novos arquivos?"
            )
            if not continuar:
                break
