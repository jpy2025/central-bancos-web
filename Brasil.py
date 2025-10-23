import pdfplumber
import pandas as pd
import os
import time
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from main import LoaderDialog


# ──────────────────────────────────────────────────────────────────────────────
# 🔹 Função para versão Streamlit (Web)
# ──────────────────────────────────────────────────────────────────────────────
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    """
    Função compatível com a Central de Bancos Web (Streamlit)
    ---------------------------------------------------------
    - files: lista de caminhos de PDFs enviados
    - output_dir: pasta onde salvar os resultados
    - progress_cb: função de callback para progresso (0–100)
    - log_cb: função de callback para logs de status
    ---------------------------------------------------------
    """
    log_cb("Iniciando processamento dos arquivos do Banco do Brasil...")

    total = len(files)
    todos_dados = []

    for i, pdf_path in enumerate(files, start=1):
        nome = os.path.basename(pdf_path)
        log_cb(f"📄 Lendo arquivo {i}/{total}: {nome}")
        progress_cb(int((i - 1) / total * 60))

        try:
            dados = extrair_dados_pdf(pdf_path)
            if dados:
                todos_dados.extend(dados)
                log_cb(f"✅ {len(dados)} lançamentos extraídos de {nome}")
            else:
                log_cb(f"⚠️ Nenhum lançamento encontrado em {nome}")
        except Exception as e:
            log_cb(f"❌ Erro ao processar {nome}: {str(e)}")

        time.sleep(0.3)
        progress_cb(int((i / total) * 70))

    if todos_dados:
        excel_path = os.path.join(
            output_dir, "Banco_do_Brasil_Resultados.xlsx")
        salvar_para_excel(todos_dados, excel_path)
        log_cb(f"💾 Planilha salva em: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento válido encontrado nos PDFs.")

    progress_cb(100)
    log_cb("Processamento concluído com sucesso! 🚀")


# ──────────────────────────────────────────────────────────────────────────────
# Extração dos dados do PDF
# ──────────────────────────────────────────────────────────────────────────────
def extrair_dados_pdf(pdf_path):
    dados = []
    data_atual = None
    historico_temp = ""
    documento_temp = ""
    valor_temp = ""

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tabela = page.extract_table()
            if tabela:
                for linha in tabela:
                    if linha and len(linha) >= 7:
                        data = linha[0] if linha[0] else data_atual
                        historico = linha[4] if linha[4] else ""
                        documento = linha[5] if linha[5] else ""
                        valor = linha[6] if linha[6] else valor_temp

                        if "Dt. balancete" in str(data) or "Histórico" in str(historico) or "Valor R$" in str(valor):
                            continue

                        if not linha[0] and not linha[6]:
                            if linha[4]:
                                historico_temp += " " + linha[4].strip()
                        else:
                            if data_atual and historico_temp:
                                dados.append(
                                    [data_atual, historico_temp.strip(), documento_temp, valor_temp])
                            data_atual = data
                            documento_temp = documento.strip()
                            historico_temp = historico.strip() if historico else ""
                            valor_temp = valor.strip()

        if data_atual and historico_temp:
            dados.append([data_atual, historico_temp.strip(),
                         documento_temp, valor_temp])

    dados = [linha for linha in dados if linha[1]
             not in ["Histórico", "Saldo Anterior"]]
    return dados


# ──────────────────────────────────────────────────────────────────────────────
# Salvamento e formatação do Excel
# ──────────────────────────────────────────────────────────────────────────────
def salvar_para_excel(dados, caminho_pdf):
    pasta, nome_pdf = os.path.split(caminho_pdf)
    nome_excel = os.path.splitext(nome_pdf)[0] + ".xlsx"
    caminho_excel = os.path.join(pasta, nome_excel)

    df = pd.DataFrame(
        dados, columns=["Data", "Histórico", "Documento", "Valor"])
    df = df[~df["Histórico"].str.contains("S A L D O", na=False)]
    df = df[~df["Histórico"].str.contains("Saldo Anterior", na=False)]
    df.to_excel(caminho_excel, index=False)

    wb = load_workbook(caminho_excel)
    ws = wb.active
    ws.freeze_panes = 'A2'

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
            if cell.column == 4:
                try:
                    valor_str = str(cell.value).strip()
                    if "C" in valor_str:
                        valor_str = valor_str.replace(
                            "C", "").replace(".", "").replace(",", ".")
                        valor = float(valor_str)
                        cell.value = valor
                        cell.font = Font(color="0000FF")  # Azul = positivo
                    elif "D" in valor_str:
                        valor_str = valor_str.replace(
                            "D", "").replace(".", "").replace(",", ".")
                        valor = float(valor_str)
                        cell.value = -valor
                        cell.font = Font(color="FF0000")  # Vermelho = negativo
                    else:
                        valor_str = valor_str.replace(
                            ".", "").replace(",", ".")
                        valor = float(valor_str)
                        cell.value = valor

                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                except:
                    pass
            else:
                cell.alignment = Alignment(horizontal="left")

    for col in ws.columns:
        max_len = max(len(str(cell.value))
                      if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 3

    num_linhas = ws.max_row
    ultima_coluna = get_column_letter(ws.max_column)
    tab = Table(displayName="TabelaLancamentos",
                ref=f"A1:{ultima_coluna}{num_linhas}")

    estilo = TableStyleInfo(name="TableStyleMedium2",
                            showFirstColumn=False,
                            showLastColumn=False,
                            showRowStripes=True,
                            showColumnStripes=False)
    tab.tableStyleInfo = estilo
    ws.add_table(tab)

    wb.save(caminho_excel)
    return caminho_excel


# ──────────────────────────────────────────────────────────────────────────────
# Fluxo padrão PyQt5 (Desktop)
# ──────────────────────────────────────────────────────────────────────────────
def processar_pdf_custom(qt_parent):
    try:
        while True:
            arquivos_pdf, _ = QFileDialog.getOpenFileNames(
                qt_parent, "Selecione os extratos do Banco do Brasil", "", "PDF Files (*.pdf)"
            )
            if not arquivos_pdf:
                break

            for arquivo_pdf in arquivos_pdf:
                dialog = LoaderDialog(qt_parent, qt_parent.light_theme)
                dialog.show()
                QApplication.processEvents()

                try:
                    dialog.atualizar_progresso(10)
                    dados = extrair_dados_pdf(arquivo_pdf)
                    dialog.atualizar_progresso(60)

                    if dados:
                        salvar_para_excel(dados, arquivo_pdf)
                        dialog.atualizar_progresso(90)
                    else:
                        QMessageBox.warning(
                            qt_parent, "Aviso", f"Nenhum dado encontrado no arquivo:\n{os.path.basename(arquivo_pdf)}"
                        )

                    dialog.atualizar_progresso(100)

                except Exception as e:
                    QMessageBox.critical(
                        qt_parent, "Erro", f"Erro no arquivo {os.path.basename(arquivo_pdf)}:\n{str(e)}"
                    )
                    dialog.atualizar_progresso(100)

                dialog.accept()

            continuar = qt_parent.mostrar_confirmacao(
                "Concluído", "Todos os arquivos selecionados foram processados.\n\nDeseja selecionar novos arquivos?"
            )
            if not continuar:
                break

    except Exception as erro_final:
        QMessageBox.critical(qt_parent, "Erro geral",
                             f"Erro inesperado:\n{str(erro_final)}")
