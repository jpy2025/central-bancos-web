# ==========================================================
# Módulo: Stone.py
# Compatível com:
#   - Streamlit (função processar_pdf_streamlit)
#   - PyQt5 (função processar_pdf_custom)
# ==========================================================

import os
import re
import time
import pdfplumber
import pandas as pd
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ✅ Importa a classe LoaderDialog
from main import LoaderDialog


# ==========================================================
# 🔹 Extração dos lançamentos do PDF
# ==========================================================
def extrair_dados_pdf(caminho_pdf):
    dados = []
    palavras_chave_ignorar = [
        "informações do comprovante", "código de autenticação", "ouvidoria",
        "meajuda@stone.com.br", "cnpj", "ligue para", "fale com a gente",
        "extrato de conta corrente", "emitido no dia", "titular", "instituição",
        "documento", "período", "agência", "conta", "saldo (r$)", "contraparte"
    ]

    with pdfplumber.open(caminho_pdf) as pdf:
        linhas = []
        for pagina in pdf.pages:
            texto = pagina.extract_text()
            if texto:
                linhas.extend(texto.split('\n'))

    i = 0
    data_atual = None
    descricao_temp = []

    while i < len(linhas):
        linha = linhas[i].strip()
        if any(p in linha.lower() for p in palavras_chave_ignorar):
            i += 1
            continue

        match_data = re.match(r'^(\d{2}/\d{2}/\d{4})', linha)
        if match_data:
            if data_atual and descricao_temp:
                descricao = ' '.join(descricao_temp).strip()
                dados.append({
                    'DATA': data_atual,
                    'LANÇAMENTO': descricao,
                    'VALOR (R$)': valor_encontrado,
                    'COR': cor
                })
                descricao_temp = []

            data_atual = match_data.group(1)
            tipo_linha = linha.lower()
            is_credito = "crédito" in tipo_linha
            cor = 'FF0000' if is_credito else '0000FF'

            valores = re.findall(r'\d{1,3}(?:\.\d{3})*,\d{2}', linha)
            valor_encontrado = float(valores[0].replace(
                '.', '').replace(',', '.')) if valores else 0.0

            descricao_temp = []
            i += 1
            while i < len(linhas):
                prox_linha = linhas[i].strip()
                if not prox_linha:
                    i += 1
                    continue
                if re.match(r'^\d{2}/\d{2}/\d{4}', prox_linha):
                    i -= 1
                    break
                if any(p in prox_linha.lower() for p in palavras_chave_ignorar):
                    break
                if re.search(r'\d{1,3}(?:\.\d{3})*,\d{2}', prox_linha):
                    break
                descricao_temp.append(prox_linha)
                i += 1
        i += 1

    if data_atual and descricao_temp:
        descricao = ' '.join(descricao_temp).strip()
        if not any(p in descricao.lower() for p in palavras_chave_ignorar):
            dados.append({
                'DATA': data_atual,
                'LANÇAMENTO': descricao,
                'VALOR (R$)': valor_encontrado,
                'COR': cor
            })

    return dados


# ==========================================================
# 🔹 Formatação e salvamento em Excel
# ==========================================================
def salvar_em_excel(dados, caminho_pdf):
    df = pd.DataFrame(dados)
    caminho_excel = caminho_pdf.replace('.pdf', '.xlsx')
    df.drop(columns=['COR']).to_excel(caminho_excel, index=False)

    wb = load_workbook(caminho_excel)
    ws = wb.active
    ws.freeze_panes = 'A2'

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for idx, row in enumerate(dados, start=2):
        valor_cell = ws[f'C{idx}']
        valor_cell.font = Font(color=row['COR'])
        valor_cell.number_format = '#,##0.00'
        valor_cell.alignment = Alignment(horizontal="right")

        for col in ('A', 'B'):
            ws[f"{col}{idx}"].border = thin_border
            ws[f"{col}{idx}"].alignment = Alignment(horizontal="left")
        valor_cell.border = thin_border

    for col in ws.columns:
        max_length = max(len(str(cell.value))
                         if cell.value else 0 for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2

    ultima_linha = len(dados) + 1
    tab = Table(displayName="TabelaStone", ref=f"A1:C{ultima_linha}")
    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True
    )
    tab.tableStyleInfo = estilo
    ws.add_table(tab)

    wb.save(caminho_excel)
    return caminho_excel


# ==========================================================
# 💻 Modo Desktop (PyQt5)
# ==========================================================
def processar_pdf_custom(janela):
    while True:
        arquivos, _ = QFileDialog.getOpenFileNames(
            janela, "Selecione um ou mais extratos da Stone", "", "Arquivos PDF (*.pdf)"
        )
        if not arquivos:
            break

        for caminho_pdf in arquivos:
            dialog = LoaderDialog(janela, janela.light_theme)
            dialog.show()
            QApplication.processEvents()

            try:
                dialog.atualizar_progresso(10)

                dados = extrair_dados_pdf(caminho_pdf)
                dialog.atualizar_progresso(50)

                if not dados:
                    QMessageBox.warning(
                        janela, "Aviso", f"Nenhum dado válido foi encontrado no arquivo:\n{caminho_pdf}")
                    dialog.atualizar_progresso(100)
                    dialog.accept()
                    continue

                salvar_em_excel(dados, caminho_pdf)
                dialog.atualizar_progresso(100)

            except Exception as e:
                QMessageBox.critical(janela, "Erro", f"Ocorreu um erro:\n{e}")
                dialog.atualizar_progresso(100)

            dialog.accept()

        continuar = janela.mostrar_confirmacao(
            "Concluído",
            "Todos os arquivos selecionados foram processados.\n\nDeseja selecionar novos arquivos?"
        )
        if not continuar:
            break


# ==========================================================
# 🌐 Modo Web (Streamlit)
# ==========================================================
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    log_cb("Iniciando processamento de extratos Stone...")

    total = len(files)
    registros = []

    for i, pdf_path in enumerate(files, start=1):
        log_cb(f"Lendo arquivo {i}/{total}: {os.path.basename(pdf_path)}")
        time.sleep(0.3)
        progress_cb(int((i / total) * 70))

        try:
            dados = extrair_dados_pdf(pdf_path)
            if not dados:
                log_cb(
                    f"⚠️ Nenhum lançamento encontrado em {os.path.basename(pdf_path)}")
                continue

            registros.extend(dados)

        except Exception as e:
            log_cb(f"❌ Erro ao processar {os.path.basename(pdf_path)}: {e}")

    if registros:
        df_final = pd.DataFrame(registros)
        os.makedirs(output_dir, exist_ok=True)
        excel_path = os.path.join(output_dir, "Stone_Resultados.xlsx")
        df_final.drop(columns=['COR'], errors='ignore').to_excel(
            excel_path, index=False)
        log_cb(f"✅ Planilha gerada: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento encontrado nos PDFs enviados.")

    progress_cb(100)
    log_cb("Processamento concluído com sucesso! 🚀")
    return True
