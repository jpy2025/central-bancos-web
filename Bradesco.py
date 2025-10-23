import os
import re
import fitz
import pandas as pd

from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication, QDialog
from PyQt5.QtCore import QThread, pyqtSignal, QObject

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from main import LoaderDialog
import time


# ──────────────────────────────────────────────────────────────────────────────
# 🔹 Função para versão Streamlit (Web)
# ──────────────────────────────────────────────────────────────────────────────
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    """
    Função compatível com a Central de Bancos Web (Streamlit)
    ---------------------------------------------------------
    - files: lista de caminhos de PDFs enviados
    - output_dir: pasta onde salvar os resultados
    - progress_cb: função de callback para progresso (0–100)
    - log_cb: função de callback para logs de status
    ---------------------------------------------------------
    """
    log_cb("Iniciando processamento dos arquivos do Bradesco...")

    total = len(files)
    todos_dados = []

    for i, pdf_path in enumerate(files, start=1):
        nome = os.path.basename(pdf_path)
        log_cb(f"📄 Lendo arquivo {i}/{total}: {nome}")
        progress_cb(int((i - 1) / total * 60))

        try:
            dados = extrair_lancamentos(pdf_path)
            if dados:
                todos_dados.extend(dados)
                log_cb(f"✅ {len(dados)} lançamentos extraídos de {nome}")
            else:
                log_cb(f"⚠️ Nenhum lançamento encontrado em {nome}")
        except Exception as e:
            log_cb(f"❌ Erro ao processar {nome}: {str(e)}")

        time.sleep(0.3)
        progress_cb(int((i / total) * 70))

    if todos_dados:
        excel_path = os.path.join(output_dir, "Bradesco_Resultados.xlsx")
        salvar_excel(todos_dados, excel_path)
        log_cb(f"💾 Planilha salva em: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento válido encontrado nos PDFs.")

    progress_cb(100)
    log_cb("Processamento concluído com sucesso! 🚀")


# ──────────────────────────────────────────────────────────────────────────────
# Worker: executa extração + salvamento em background e emite marcos (10/60/90)
# ──────────────────────────────────────────────────────────────────────────────
class FileWorker(QThread):
    milestone = pyqtSignal(int)                 # 10, 60, 90
    finished_with_status = pyqtSignal(str, str)  # ("ok"|"vazio"|"erro", extra)

    def __init__(self, caminho_pdf: str, parent=None):
        super().__init__(parent)
        self._caminho_pdf = caminho_pdf

    def run(self):
        try:
            self.milestone.emit(10)
            dados = extrair_lancamentos(self._caminho_pdf)
            self.milestone.emit(60)

            if not dados:
                self.finished_with_status.emit(
                    "vazio", os.path.basename(self._caminho_pdf))
                return

            salvar_excel(dados, self._caminho_pdf)
            self.milestone.emit(90)
            self.finished_with_status.emit("ok", "")
        except Exception as e:
            self.finished_with_status.emit(
                "erro", f"{os.path.basename(self._caminho_pdf)}: {str(e)}")


# ──────────────────────────────────────────────────────────────────────────────
# Loader suave: anima de 5 em 5% até o alvo atual; segura no alvo até mudar
# ──────────────────────────────────────────────────────────────────────────────
class ProgressThread(QThread):
    progress = pyqtSignal(int)

    def __init__(self, step_ms: int = 120, parent=None):
        super().__init__(parent)
        self._step_ms = max(16, int(step_ms))
        self._val = 0
        self._target = 0
        self._finishing = False

    def set_target(self, percent: int):
        self._target = max(0, min(int(percent), 100))

    def stop_and_finish(self):
        self._finishing = True
        self._target = 100

    def run(self):
        self._val = 0
        self._target = 0
        self.progress.emit(0)
        while True:
            if self._val >= self._target:
                if self._finishing and self._val >= 100:
                    self.progress.emit(100)
                    break
                self.msleep(self._step_ms)
                continue
            self._val = min(self._target, self._val + 5)
            self.progress.emit(self._val)
            self.msleep(self._step_ms)


# ──────────────────────────────────────────────────────────────────────────────
# Filtros de texto e extração (fitz)
# ──────────────────────────────────────────────────────────────────────────────
def linha_eh_cabecalho_ou_rodape(linha: str) -> bool:
    padroes_excluir = [
        r"Folha \d+/\d+",
        r"Extrato Mensal\s*/\s*Por Período",
        r"CNPJ[:\s]*\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}",
        r"Nome do usuário:.*",
        r"Data da operação: \d{2}/\d{2}/\d{4}",
        r"UNISESCAP CEARA",
        r"bradesco",
        r"net empresa",
        r"Assistente de IA",
    ]
    return any(re.search(p, linha, re.IGNORECASE) for p in padroes_excluir)


def extrair_lancamentos(pdf_path: str):
    texto_total = ""
    with fitz.open(pdf_path) as doc:
        for pagina in doc:
            texto_total += pagina.get_text()

    linhas = texto_total.splitlines()
    data_regex = re.compile(r"^\d{2}/\d{2}/\d{4}$")
    valor_regex = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}")
    ignorar_regex = re.compile(r"\b(SALDO|TOTAL)\b", re.IGNORECASE)

    lancamentos = []
    data_atual = None
    descricao_temp = ""
    i = 0

    while i < len(linhas):
        linha = linhas[i].strip()

        if linha_eh_cabecalho_ou_rodape(linha):
            i += 1
            continue

        if data_regex.match(linha):
            data_atual = linha
            descricao_temp = ""
            i += 1
            continue

        if not data_atual:
            i += 1
            continue

        if ignorar_regex.search(linha):
            i += 1
            continue

        valores = re.findall(valor_regex, linha)
        valores_float = [float(v.replace('.', '').replace(',', '.'))
                         for v in valores]

        if valores_float:
            if descricao_temp:
                descricao = descricao_temp.strip()
                if not ignorar_regex.search(descricao):
                    for valor in valores_float:
                        lancamentos.append(
                            {'Data': data_atual, 'Lançamento': descricao, 'Valor (R$)': valor})
                descricao_temp = ""
            else:
                if (i > 0 and not data_regex.match(linhas[i - 1].strip()) and not valor_regex.search(linhas[i - 1])):
                    descricao_temp = linhas[i - 1].strip()
                    if not ignorar_regex.search(descricao_temp):
                        for valor in valores_float:
                            lancamentos.append(
                                {'Data': data_atual, 'Lançamento': descricao_temp, 'Valor (R$)': valor})
                    descricao_temp = ""
        else:
            if linha and not ignorar_regex.search(linha):
                descricao_temp = (descricao_temp + " " +
                                  linha).strip() if descricao_temp else linha

        i += 1

    return lancamentos


def salvar_excel(dados, caminho_pdf: str) -> str:
    df = pd.DataFrame(dados)

    pasta = os.path.dirname(caminho_pdf)
    nome_base = os.path.splitext(os.path.basename(caminho_pdf))[0]
    excel_path = os.path.join(pasta, f"{nome_base}.xlsx")

    df.to_excel(excel_path, index=False)

    wb = load_workbook(excel_path)
    ws = wb.active
    ws.freeze_panes = 'A2'

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    valor_col_index = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == 'Valor (R$)':
            valor_col_index = idx
            break

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.column == valor_col_index:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                    if cell.value > 0:
                        cell.font = Font(color="0000FF")
                    elif cell.value < 0:
                        cell.font = Font(color="FF0000")
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    for col in ws.columns:
        max_len = max(len(str(cell.value))
                      if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 2

    num_linhas = ws.max_row
    tab = Table(displayName="TabelaLancamentos", ref=f"A1:C{num_linhas}")
    estilo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = estilo
    ws.add_table(tab)

    wb.save(excel_path)
    return excel_path


def _fechar_loader_com_seguranca(dialog: QDialog):
    try:
        if isinstance(dialog, QDialog):
            dialog.accept()
        else:
            dialog.close()
    except Exception:
        dialog.close()
    dialog.deleteLater()
    QApplication.processEvents()


def processar_pdf_custom(qt_parent):
    try:
        while True:
            caminhos_pdf, _ = QFileDialog.getOpenFileNames(
                qt_parent, "Selecione um ou mais extratos do Bradesco", "", "PDF Files (*.pdf)"
            )
            if not caminhos_pdf:
                break

            for caminho_pdf in caminhos_pdf:
                dialog = LoaderDialog(qt_parent, getattr(
                    qt_parent, "light_theme", True))
                dialog.setWindowTitle("Carregando arquivo, aguarde...")
                dialog.atualizar_progresso(0)
                dialog.show()
                QApplication.processEvents()

                loader = ProgressThread(step_ms=120, parent=dialog)
                loader.progress.connect(dialog.atualizar_progresso)
                loader.start()

                worker = FileWorker(caminho_pdf, parent=dialog)
                worker.milestone.connect(loader.set_target)

                status_result = {"status": None, "extra": ""}

                def _on_finished(status, extra):
                    status_result["status"] = status
                    status_result["extra"] = extra

                worker.finished_with_status.connect(_on_finished)
                worker.start()

                while worker.isRunning():
                    QApplication.processEvents()
                    QThread.msleep(50)

                loader.stop_and_finish()
                loader.wait()
                dialog.atualizar_progresso(100)
                QApplication.processEvents()

                st, ex = status_result["status"], status_result["extra"]
                if st == "vazio":
                    QMessageBox.warning(
                        qt_parent, "Aviso", f"Nenhum lançamento encontrado no arquivo:\n{ex}")
                elif st == "erro":
                    QMessageBox.critical(
                        qt_parent, "Erro", f"Erro no arquivo {ex}")

                _fechar_loader_com_seguranca(dialog)

            continuar = qt_parent.mostrar_confirmacao(
                "Concluído", "Todos os arquivos selecionados foram processados.\n\nDeseja selecionar novos arquivos?"
            )
            if not continuar:
                break

    except Exception as erro_final:
        QMessageBox.critical(qt_parent, "Erro geral",
                             f"Erro inesperado:\n{str(erro_final)}")
