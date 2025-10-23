import re
import os
import time
import pandas as pd
import fitz  # PyMuPDF
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ✅ Importa LoaderDialog
from main import LoaderDialog


# ──────────────────────────────────────────────────────────────────────────────
# 🔹 Função compatível com a versão Web (Streamlit)
# ──────────────────────────────────────────────────────────────────────────────
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    """
    Compatível com a Central de Bancos Web (Streamlit)
    ---------------------------------------------------------
    - files: lista de PDFs enviados
    - output_dir: pasta onde salvar o Excel
    - progress_cb: callback (0–100)
    - log_cb: função de log de mensagens
    ---------------------------------------------------------
    """
    log_cb("Iniciando processamento dos extratos da Caixa Econômica Federal...")

    total = len(files)
    dfs = []

    for i, pdf_path in enumerate(files, start=1):
        nome = os.path.basename(pdf_path)
        log_cb(f"📄 Processando arquivo {i}/{total}: {nome}")
        progress_cb(int((i - 1) / total * 60))

        try:
            df = extrair_lancamentos(pdf_path)
            if not df.empty:
                dfs.append(df)
                log_cb(f"✅ {len(df)} lançamentos extraídos de {nome}")
            else:
                log_cb(f"⚠️ Nenhum lançamento encontrado em {nome}")
        except Exception as e:
            log_cb(f"❌ Erro ao processar {nome}: {e}")

        time.sleep(0.2)
        progress_cb(int((i / total) * 80))

    if dfs:
        df_final = pd.concat(dfs, ignore_index=True)
        excel_path = os.path.join(output_dir, "Caixa_Resultados.xlsx")

        df_final.to_excel(excel_path, index=False, columns=[
            "Data Mov.", "Histórico", "Valor"
        ])

        aplicar_formatacao_excel(excel_path, df_final["Tipo"].tolist())
        log_cb(f"💾 Planilha final salva em: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento válido encontrado.")

    progress_cb(100)
    log_cb("✅ Processamento concluído com sucesso! 🚀")


# ──────────────────────────────────────────────────────────────────────────────
# Extração de lançamentos do PDF
# ──────────────────────────────────────────────────────────────────────────────
def extrair_lancamentos(pdf_path):
    doc = fitz.open(pdf_path)
    texto = ""
    for pagina in doc:
        texto += pagina.get_text()
    doc.close()

    padrao = re.findall(
        r'(\d{2}/\d{2}/\d{4})\s+\d{6}\s+(.*?)\s+([\d.,]+)\s+([CD])', texto)

    dados = []
    for data, historico, valor, tipo in padrao:
        valor_float = float(valor.replace('.', '').replace(',', '.'))
        dados.append([data, historico.strip(), valor_float, tipo])

    df = pd.DataFrame(
        dados, columns=["Data Mov.", "Histórico", "Valor", "Tipo"])
    df = df[~df["Histórico"].str.upper().str.contains("SALDO")]
    return df


# ──────────────────────────────────────────────────────────────────────────────
# Formatação do Excel
# ──────────────────────────────────────────────────────────────────────────────
def aplicar_formatacao_excel(excel_path, tipos):
    wb = load_workbook(excel_path)
    ws = wb.active

    ws.freeze_panes = 'A2'

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    # Cabeçalhos
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Colore e formata valores
    for idx, row in enumerate(ws.iter_rows(min_row=2, min_col=3, max_col=3), start=0):
        cell = row[0]
        cell.number_format = '#,##0.00'
        cell.border = thin_border
        if tipos[idx] == 'C':
            cell.font = Font(color="0000FF")  # Azul = crédito
        else:
            cell.font = Font(color="FF0000")  # Vermelho = débito
        cell.alignment = Alignment(horizontal="right")

    # Alinha as demais colunas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            if cell.col_idx != 3:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left")

    # Ajusta largura das colunas
    for col in ws.columns:
        max_len = max(len(str(cell.value))
                      if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 2

    # Adiciona tabela
    num_linhas = ws.max_row
    tab = Table(displayName="TabelaLancamentos", ref=f"A1:C{num_linhas}")
    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = estilo
    ws.add_table(tab)

    wb.save(excel_path)


# ──────────────────────────────────────────────────────────────────────────────
# Fluxo padrão PyQt5 (Desktop)
# ──────────────────────────────────────────────────────────────────────────────
def processar_pdf_custom(qt_parent):
    while True:
        arquivos_pdf, _ = QFileDialog.getOpenFileNames(
            qt_parent, "Selecione um ou mais PDFs da Caixa", "", "PDF Files (*.pdf)"
        )
        if not arquivos_pdf:
            break

        for pdf_path in arquivos_pdf:
            dialog = LoaderDialog(qt_parent, qt_parent.light_theme)
            dialog.show()
            QApplication.processEvents()

            try:
                dialog.atualizar_progresso(10)
                df = extrair_lancamentos(pdf_path)
                dialog.atualizar_progresso(50)

                if df.empty:
                    QMessageBox.warning(
                        qt_parent, "Aviso", f"Nenhum lançamento encontrado no arquivo:\n{os.path.basename(pdf_path)}")
                    dialog.atualizar_progresso(100)
                    dialog.accept()
                    continue

                excel_path = os.path.splitext(pdf_path)[0] + ".xlsx"
                df.to_excel(excel_path, index=False, columns=[
                    "Data Mov.", "Histórico", "Valor"])
                dialog.atualizar_progresso(80)

                aplicar_formatacao_excel(excel_path, df["Tipo"].tolist())
                dialog.atualizar_progresso(100)

            except Exception as e:
                QMessageBox.critical(
                    qt_parent, "Erro", f"Ocorreu um erro no arquivo {os.path.basename(pdf_path)}:\n{e}")
                dialog.atualizar_progresso(100)

            dialog.accept()

            continuar = qt_parent.mostrar_confirmacao(
                "Concluído",
                "Todos os arquivos selecionados foram processados.\n\nDeseja selecionar novos arquivos?"
            )
            if not continuar:
                break
