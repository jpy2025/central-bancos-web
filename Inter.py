import os
import re
import time
import pandas as pd
import fitz  # PyMuPDF
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ✅ Importa a classe LoaderDialog
from main import LoaderDialog


# ──────────────────────────────────────────────────────────────────────────────
# 🔹 Função compatível com o modo Web (Streamlit)
# ──────────────────────────────────────────────────────────────────────────────
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    """
    Compatível com Central de Bancos Web (Streamlit)
    ---------------------------------------------------------
    - files: lista de PDFs enviados
    - output_dir: pasta onde salvar o Excel
    - progress_cb: callback (0–100)
    - log_cb: função de log de mensagens
    ---------------------------------------------------------
    """
    log_cb("Iniciando processamento dos extratos do Banco Inter...")
    todos_dados = []
    total = len(files)

    for i, pdf_path in enumerate(files, start=1):
        nome = os.path.basename(pdf_path)
        log_cb(f"📄 Processando {i}/{total}: {nome}")
        progress_cb(int((i - 1) / total * 60))

        try:
            df = extrair_lancamentos_por_posicao(pdf_path)
            if not df.empty:
                todos_dados.append(df)
                log_cb(f"✅ {len(df)} lançamentos extraídos de {nome}")
            else:
                log_cb(f"⚠️ Nenhum lançamento válido em {nome}")
        except Exception as e:
            log_cb(f"❌ Erro ao processar {nome}: {str(e)}")

        time.sleep(0.2)
        progress_cb(int((i / total) * 80))

    if todos_dados:
        df_final = pd.concat(todos_dados, ignore_index=True)
        excel_path = os.path.join(output_dir, "Inter_Resultados.xlsx")

        df_final.to_excel(excel_path, index=False)
        aplicar_formatacao_excel(excel_path)
        log_cb(f"💾 Planilha salva em: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento válido encontrado nos PDFs.")

    progress_cb(100)
    log_cb("✅ Processamento concluído com sucesso! 🚀")


# ──────────────────────────────────────────────────────────────────────────────
# Função auxiliar: converte datas por extenso (ex: '5 de janeiro de 2025')
# ──────────────────────────────────────────────────────────────────────────────
def formatar_data(data_extenso):
    meses = {
        'janeiro': '01', 'fevereiro': '02', 'março': '03', 'abril': '04',
        'maio': '05', 'junho': '06', 'julho': '07', 'agosto': '08',
        'setembro': '09', 'outubro': '10', 'novembro': '11', 'dezembro': '12'
    }
    match = re.search(r'(\d{1,2}) de (\w+) de (\d{4})', data_extenso.lower())
    if match:
        dia, mes_nome, ano = match.groups()
        dia = dia.zfill(2)
        mes = meses.get(mes_nome, '??')
        return f"{dia}/{mes}/{ano}"
    return None


# ──────────────────────────────────────────────────────────────────────────────
# Extração dos lançamentos por posição (BTG / Inter usam blocos de texto)
# ──────────────────────────────────────────────────────────────────────────────
def extrair_lancamentos_por_posicao(pdf_path):
    doc = fitz.open(pdf_path)
    dados = []
    data_atual = ""

    for pagina in doc:
        blocos = pagina.get_text("blocks")
        blocos.sort(key=lambda b: (round(b[1]), b[0]))

        linhas = {}
        for b in blocos:
            x0, y0, x1, y1, texto = b[:5]
            y_key = round(y0)
            if y_key not in linhas:
                linhas[y_key] = []
            linhas[y_key].append((x0, texto.strip()))

        for y in sorted(linhas.keys()):
            linha = linhas[y]
            linha.sort(key=lambda x: x[0])
            textos = [t for _, t in linha]
            linha_texto = " ".join(textos)

            # Detectar data por extenso (ex: '5 de janeiro de 2025')
            data_detectada = re.search(r'\d{1,2} de \w+ de \d{4}', linha_texto)
            if data_detectada:
                data_atual = formatar_data(data_detectada.group())
                continue

            if not data_atual or 'R$' not in linha_texto:
                continue

            valores = re.findall(r'-?R\$\s?[\d\.]+,\d{2}', linha_texto)
            if not valores:
                continue

            valor_raw = valores[0]
            valor_limpo = valor_raw.replace('R$', '').replace(
                ' ', '').replace('.', '').replace(',', '.')

            try:
                valor_float = round(float(valor_limpo), 2)
                if '-' in valor_raw:
                    valor_float = -abs(valor_float)

                historico = linha_texto.split(valor_raw)[0].strip()
                dados.append({
                    "Data": data_atual,
                    "Histórico": historico,
                    "Valor": valor_float
                })
            except ValueError:
                continue

    doc.close()
    return pd.DataFrame(dados)


# ──────────────────────────────────────────────────────────────────────────────
# Aplicar formatação ao Excel (padrão azul e vermelho)
# ──────────────────────────────────────────────────────────────────────────────
def aplicar_formatacao_excel(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.col_idx == 3:  # Coluna Valor
                cell.font = Font(color="FF0000") if cell.value < 0 else Font(
                    color="0000FF")
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    for col in ws.columns:
        max_len = max(len(str(cell.value))
                      if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 2

    num_linhas = ws.max_row
    tab = Table(displayName="TabelaLancamentos", ref=f"A1:C{num_linhas}")
    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = estilo
    ws.add_table(tab)
    wb.save(excel_path)


# ──────────────────────────────────────────────────────────────────────────────
# Fluxo padrão PyQt5 (Desktop)
# ──────────────────────────────────────────────────────────────────────────────
def processar_pdf_custom(parent_widget):
    while True:
        arquivos, _ = QFileDialog.getOpenFileNames(
            parent_widget, "Selecione os extratos do Inter", "", "PDF Files (*.pdf)"
        )

        if not arquivos:
            break

        for file_path in arquivos:
            dialog = LoaderDialog(parent_widget, parent_widget.light_theme)
            dialog.show()
            QApplication.processEvents()

            try:
                dialog.atualizar_progresso(10)
                df = extrair_lancamentos_por_posicao(file_path)
                dialog.atualizar_progresso(60)

                if df.empty:
                    QMessageBox.warning(
                        parent_widget, "Erro", f"Nenhum lançamento válido em:\n{os.path.basename(file_path)}."
                    )
                    dialog.atualizar_progresso(100)
                    dialog.accept()
                    continue

                df.to_excel(file_path.replace(".pdf", ".xlsx"), index=False)
                aplicar_formatacao_excel(file_path.replace(".pdf", ".xlsx"))
                dialog.atualizar_progresso(100)

            except Exception as e:
                QMessageBox.critical(
                    parent_widget, "Erro", f"Ocorreu um erro no arquivo {os.path.basename(file_path)}:\n{e}"
                )
                dialog.atualizar_progresso(100)

            dialog.accept()

            continuar = parent_widget.mostrar_confirmacao(
                "Concluído",
                "Todos os arquivos selecionados foram processados.\n\nDeseja selecionar novos arquivos?"
            )
            if not continuar:
                break
