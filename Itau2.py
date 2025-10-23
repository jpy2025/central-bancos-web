import os
import re
import fitz  # PyMuPDF
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from main import LoaderDialog


# ══════════════════════════════════════════════════════════════════════════════
# 🔹 Versão para Central Web (Streamlit)
# ══════════════════════════════════════════════════════════════════════════════
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    """
    Compatível com Central de Bancos Web (Streamlit)
    ---------------------------------------------------------
    - files: lista de PDFs enviados
    - output_dir: pasta onde salvar o Excel
    - progress_cb: callback (0–100)
    - log_cb: função de log
    ---------------------------------------------------------
    """
    log_cb("Iniciando processamento de extratos (formato Daycoval simples)...")
    total = len(files)
    todos_dfs = []

    for i, caminho_pdf in enumerate(files, start=1):
        nome = os.path.basename(caminho_pdf)
        log_cb(f"📄 Processando {i}/{total}: {nome}")
        progress_cb(int((i - 1) / total * 60))

        try:
            df = extrair_lancamentos_pdf(caminho_pdf)
            if not df.empty:
                todos_dfs.append(df)
                log_cb(f"✅ {len(df)} lançamentos extraídos de {nome}")
            else:
                log_cb(f"⚠️ Nenhum lançamento encontrado em {nome}")
        except Exception as e:
            log_cb(f"❌ Erro ao processar {nome}: {str(e)}")

    if todos_dfs:
        df_final = pd.concat(todos_dfs, ignore_index=True)
        excel_path = os.path.join(output_dir, "Extratos_Resultados.xlsx")
        salvar_em_excel(df_final, excel_path)
        log_cb(f"💾 Planilha salva em: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento válido encontrado nos PDFs.")

    progress_cb(100)
    log_cb("✅ Processamento concluído com sucesso! 🚀")


# ══════════════════════════════════════════════════════════════════════════════
# Extração de lançamentos (texto por blocos)
# ══════════════════════════════════════════════════════════════════════════════
def extrair_lancamentos_pdf(caminho_pdf):
    doc = fitz.open(caminho_pdf)
    linhas_extraidas = []

    for pagina in doc:
        blocks = pagina.get_text("blocks")
        for b in blocks:
            for linha in b[4].split('\n'):
                linha_limpa = linha.strip()
                if linha_limpa:
                    linhas_extraidas.append(linha_limpa)

    doc.close()

    blocos = []
    i = 0
    while i < len(linhas_extraidas):
        linha = linhas_extraidas[i]
        if re.match(r"\d{2}/\d{2}/\d{4}", linha):
            data = linha
            i += 1
            texto = []
            valor = None

            while i < len(linhas_extraidas):
                atual = linhas_extraidas[i].strip()

                # Detecta fim de bloco (valor monetário)
                if re.match(r"^-?\d{1,3}(?:\.\d{3})*,\d{2}$", atual) or re.match(r"^-?\d+,\d{2}$", atual):
                    try:
                        valor = float(atual.replace(".", "").replace(",", "."))
                        i += 1
                        break
                    except ValueError:
                        pass

                texto.append(atual)
                i += 1

            descricao = " ".join(texto)

            # Ignora saldos e cabeçalhos
            if any(p in descricao.lower() for p in ["saldo", "sdo anterior", "total", "limite"]):
                continue

            blocos.append((data, descricao.strip(), valor))
        else:
            i += 1

    df = pd.DataFrame(blocos, columns=["Data", "Lançamento", "Valor (R$)"])
    df["Data"] = pd.to_datetime(df["Data"], format="%d/%m/%Y", errors='coerce')
    df = df.dropna(subset=["Data"]).reset_index(drop=True)
    df["Data"] = df["Data"].dt.strftime("%d/%m/%Y")

    return df


# ══════════════════════════════════════════════════════════════════════════════
# Criação e formatação do Excel
# ══════════════════════════════════════════════════════════════════════════════
def salvar_em_excel(df, caminho_excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lançamentos"

    bold = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")
    money_fmt = "#,##0.00"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.freeze_panes = 'A2'

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if r_idx == 1:
                cell.font = bold
                cell.alignment = center
                cell.fill = header_fill
            elif c_idx == 3:
                cell.number_format = money_fmt
                if isinstance(value, float):
                    if value < 0:
                        cell.font = Font(color="FF0000")  # Vermelho (negativo)
                    else:
                        cell.font = Font(color="0000FF")  # Azul (positivo)
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Ajuste de colunas
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 16

    wb.save(caminho_excel)


# ══════════════════════════════════════════════════════════════════════════════
# Fluxo PyQt5 (Desktop)
# ══════════════════════════════════════════════════════════════════════════════
def processar_pdf_custom(janela):
    while True:
        arquivos, _ = QFileDialog.getOpenFileNames(
            janela,
            "Selecione um ou mais extratos PDF",
            "",
            "Arquivos PDF (*.pdf)"
        )

        if not arquivos:
            break

        for caminho_pdf in arquivos:
            dialog = LoaderDialog(janela, janela.light_theme)
            dialog.show()
            QApplication.processEvents()

            try:
                dialog.atualizar_progresso(10)
                df = extrair_lancamentos_pdf(caminho_pdf)
                dialog.atualizar_progresso(60)

                caminho_excel = Path(caminho_pdf).with_suffix('.xlsx')
                salvar_em_excel(df, caminho_excel)
                dialog.atualizar_progresso(100)

            except Exception as e:
                QMessageBox.critical(
                    janela, "Erro", f"Erro ao processar:\n{os.path.basename(caminho_pdf)}\n\n{str(e)}"
                )

            dialog.accept()

        continuar = janela.mostrar_confirmacao(
            "Concluído",
            "Todos os arquivos foram processados.\n\nDeseja selecionar outros arquivos?"
        )
        if not continuar:
            break
