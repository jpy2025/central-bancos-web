import os
import re
import pandas as pd
from pathlib import Path
from PyPDF2 import PdfReader
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ✅ Importa a classe LoaderDialog
from main import LoaderDialog


# ══════════════════════════════════════════════════════════════════════════════
# 🔹 Função compatível com Central Web (Streamlit)
# ══════════════════════════════════════════════════════════════════════════════
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    """
    Compatível com Central de Bancos Web (Streamlit)
    ---------------------------------------------------------
    - files: lista de PDFs enviados
    - output_dir: pasta onde salvar o Excel
    - progress_cb: callback (0–100)
    - log_cb: função de log
    ---------------------------------------------------------
    """
    log_cb("Iniciando processamento dos extratos PDF...")
    total = len(files)
    dfs = []

    for i, caminho_pdf in enumerate(files, start=1):
        nome = os.path.basename(caminho_pdf)
        log_cb(f"📄 Processando {i}/{total}: {nome}")
        progress_cb(int((i - 1) / total * 60))

        try:
            df = extrair_lancamentos_pdf(caminho_pdf)
            if not df.empty:
                dfs.append(df)
                log_cb(f"✅ {len(df)} lançamentos extraídos de {nome}")
            else:
                log_cb(f"⚠️ Nenhum lançamento encontrado em {nome}")
        except Exception as e:
            log_cb(f"❌ Erro ao processar {nome}: {e}")

    if dfs:
        df_final = pd.concat(dfs, ignore_index=True)
        excel_path = os.path.join(output_dir, "Extratos_Resultados.xlsx")
        df_final.to_excel(excel_path, index=False)
        formatar_excel(excel_path)
        log_cb(f"💾 Planilha salva em: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento válido encontrado.")

    progress_cb(100)
    log_cb("✅ Processamento concluído com sucesso! 🚀")


# ══════════════════════════════════════════════════════════════════════════════
# Extração dos lançamentos de PDF (dois modelos)
# ══════════════════════════════════════════════════════════════════════════════
def extrair_lancamentos_pdf(caminho_pdf):
    leitor = PdfReader(caminho_pdf)
    texto = ""
    for pagina in leitor.pages:
        texto += pagina.extract_text() + "\n"

    linhas = texto.splitlines()
    padrao_data_completa = re.compile(r"\d{2}/\d{2}/\d{4}")
    padrao_valor = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}")

    lancamentos = []

    # Detecta se é um modelo tabular (com cabeçalho)
    eh_modelo_tabela = any(
        "Data" in linha and "Lançamentos" in linha and "Valor" in linha
        for linha in linhas
    )

    if eh_modelo_tabela:
        # Modelo novo (tabela visível no PDF)
        for linha in linhas:
            if not padrao_data_completa.search(linha):
                continue

            try:
                valores = padrao_valor.findall(linha)
                if not valores:
                    continue
                valor = valores[-2] if len(valores) > 1 else valores[0]
                valor_float = float(valor.replace('.', '').replace(',', '.'))

                data = padrao_data_completa.search(linha).group(0)
                pos_valor = linha.find(valor)
                descricao = linha[10:pos_valor].strip()

                if not re.search(r"[a-zA-Z]", descricao):
                    continue

                lancamentos.append([data, descricao.title(), valor_float])
            except Exception:
                continue
    else:
        # Modelo antigo (texto corrido)
        padrao_data = r"(\d{2})\s*/\s*(\w{3})"
        meses = {
            "jan": "01", "fev": "02", "mar": "03", "abr": "04",
            "mai": "05", "jun": "06", "jul": "07", "ago": "08",
            "set": "09", "out": "10", "nov": "11", "dez": "12"
        }

        ano = None
        capturar = False
        data_atual = None

        for linha in linhas:
            linha_baixa = linha.lower().strip()

            # Ignora saldos
            if ("saldo" in linha_baixa or "sdo" in linha_baixa) and padrao_valor.search(linha):
                continue

            # Identifica ano
            if any(mes in linha_baixa for mes in meses) and re.search(r"\b\d{4}\b", linha_baixa):
                partes = linha_baixa.split()
                for parte in partes:
                    if parte.isdigit() and len(parte) == 4:
                        ano_int = int(parte)
                        if 2000 <= ano_int <= 2100:
                            ano = parte
                            break

            if "lançamentos período" in linha_baixa:
                capturar = True
                continue
            if not capturar:
                continue

            data_completa = padrao_data_completa.search(linha)
            if data_completa:
                data_atual = data_completa.group(0)
            else:
                data_match = re.search(padrao_data, linha)
                if not data_match or not ano:
                    continue
                dia, mes_txt = data_match.groups()
                mes_num = meses.get(mes_txt.lower().replace('#', ''), "01")
                data_atual = f"{dia.zfill(2)}/{mes_num}/{ano}"

            valor_match = padrao_valor.findall(linha)
            if len(valor_match) != 1:
                continue

            valor_str = valor_match[0]
            pos_valor = linha.find(valor_str)
            descricao = linha[:pos_valor].strip()

            if not re.search(r"[a-zA-Z]", descricao):
                continue

            lancamentos.append([
                data_atual,
                descricao.title(),
                float(valor_str.replace('.', '').replace(',', '.'))
            ])

    df = pd.DataFrame(lancamentos, columns=[
                      "Data", "Lançamento", "Valor (R$)"])
    df = df[~df["Lançamento"].str.upper().str.contains("SALDO", na=False)]
    return df


# ══════════════════════════════════════════════════════════════════════════════
# Formatação do Excel (azul, bordas, cores de valor)
# ══════════════════════════════════════════════════════════════════════════════
def formatar_excel(caminho_excel):
    wb = load_workbook(caminho_excel)
    ws = wb.active

    ws.freeze_panes = 'A2'

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.col_idx == 3:  # Coluna 'Valor (R$)'
                cell.number_format = '#,##0.00'
                cell.font = Font(color="FF0000") if cell.value < 0 else Font(
                    color="0000FF")
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    for col in ws.columns:
        max_length = max(len(str(cell.value))
                         if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    num_linhas = ws.max_row
    tab = Table(displayName="TabelaLancamentos", ref=f"A1:C{num_linhas}")
    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = estilo
    ws.add_table(tab)

    wb.save(caminho_excel)


# ══════════════════════════════════════════════════════════════════════════════
# Fluxo PyQt5 (Desktop)
# ══════════════════════════════════════════════════════════════════════════════
def processar_pdf_custom(janela):
    while True:
        arquivos, _ = QFileDialog.getOpenFileNames(
            janela, "Selecione um ou mais extratos PDF", "", "Arquivos PDF (*.pdf)"
        )

        if not arquivos:
            break

        for caminho_pdf in arquivos:
            dialog = LoaderDialog(janela, janela.light_theme)
            dialog.show()
            QApplication.processEvents()

            try:
                dialog.atualizar_progresso(10)
                df = extrair_lancamentos_pdf(caminho_pdf)
                dialog.atualizar_progresso(60)

                if df.empty:
                    QMessageBox.warning(
                        janela, "Aviso", f"Nenhum lançamento válido em:\n{os.path.basename(caminho_pdf)}"
                    )
                    dialog.accept()
                    continue

                caminho_excel = Path(caminho_pdf).with_suffix('.xlsx')
                df.to_excel(caminho_excel, index=False)
                dialog.atualizar_progresso(80)

                formatar_excel(caminho_excel)
                dialog.atualizar_progresso(100)

            except Exception as e:
                QMessageBox.critical(
                    janela, "Erro", f"Erro ao processar:\n{caminho_pdf}\n\n{str(e)}"
                )
                dialog.atualizar_progresso(100)

            dialog.accept()

        continuar = janela.mostrar_confirmacao(
            "Concluído",
            "Todos os arquivos selecionados foram processados.\n\nDeseja selecionar novos arquivos?"
        )
        if not continuar:
            break
