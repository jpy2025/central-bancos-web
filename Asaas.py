# ==========================================================
# Módulo: Asaas.py
# Compatível com:
#   - Streamlit (função processar_pdf_streamlit)
#   - PyQt5 (função processar_pdf_custom)
# ==========================================================

import fitz  # PyMuPDF
import re
import os
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path
import locale

from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from main import LoaderDialog

# Locale brasileiro
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8' if os.name !=
                 'nt' else 'Portuguese_Brazil.1252')


# ==========================================================
# 🔹 Função usada pela VERSÃO WEB (Streamlit)
# ==========================================================
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    """
    Função usada pelo app Streamlit (web-friendly)
    ---------------------------------------------------------
    files : list[str]       → Caminhos dos PDFs enviados
    output_dir : str         → Pasta onde salvar resultados
    progress_cb : callable   → Atualiza progresso (0..100)
    log_cb : callable        → Mostra mensagens na UI
    ---------------------------------------------------------
    """
    log_cb("Iniciando processamento de arquivos Asaas...")

    total = len(files)
    registros = []

    for i, pdf_path in enumerate(files, start=1):
        log_cb(f"Lendo arquivo {i}/{total}: {os.path.basename(pdf_path)}")
        time.sleep(0.4)  # simula processamento
        progress_cb(int((i / total) * 70))

        try:
            # 🔸 Extrai texto do PDF (usando PyMuPDF)
            doc = fitz.open(pdf_path)
            texto = "\n".join(page.get_text() for page in doc)
            doc.close()

            lancamentos = extrair_lancamentos(texto)

            for data, descricao, valor, cor in lancamentos:
                registros.append({
                    "Data": data,
                    "Histórico": descricao,
                    "Valor": valor,
                    "Cor": cor
                })

        except Exception as e:
            log_cb(f"❌ Erro ao processar {os.path.basename(pdf_path)}: {e}")

    # Cria o Excel de saída
    if registros:
        df = pd.DataFrame(registros)
        os.makedirs(output_dir, exist_ok=True)
        excel_path = os.path.join(output_dir, "Asaas_Resultados.xlsx")

        df.to_excel(excel_path, index=False)
        log_cb(f"✅ Planilha gerada: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento encontrado nos PDFs enviados.")

    progress_cb(100)
    log_cb("Processamento concluído com sucesso! 🚀")
    return True


# ==========================================================
# 🔹 Função auxiliar para extrair lançamentos (reutilizável)
# ==========================================================
def extrair_lancamentos(texto):
    pattern = re.compile(
        r"(\d{2}/\d{2}/\d{4})\s+(.*?)(R\$ ?-?\d[\d\.,]*)", re.DOTALL)
    matches = pattern.findall(texto)
    lancamentos = []

    for data, descricao, valor in matches:
        if 'saldo' in descricao.lower():
            continue
        descricao = re.sub(r"\s+", " ", descricao).strip()
        valor_formatado = valor.replace('.', '').replace(
            ',', '.').replace('R$', '').strip()
        valor_float = float(valor_formatado)
        cor = "0000FF" if valor_float > 0 else "FF0000"
        valor_brasileiro = f"R$ {valor_float:,.2f}".replace(
            ",", "v").replace(".", ",").replace("v", ".")
        lancamentos.append([data, descricao, valor_brasileiro, cor])

    return lancamentos


# ==========================================================
# 🔹 Função Desktop (usada pela Central de Bancos original)
# ==========================================================
def processar_pdf_custom(qt_parent):
    try:
        while True:
            caminhos_pdf, _ = QFileDialog.getOpenFileNames(
                qt_parent, "Selecione um ou mais PDFs do Asaas", "", "PDF Files (*.pdf)"
            )
            if not caminhos_pdf:
                break

            for caminho_pdf in caminhos_pdf:
                dialog = LoaderDialog(qt_parent, qt_parent.light_theme)
                dialog.show()
                QApplication.processEvents()

                try:
                    dialog.atualizar_progresso(10)
                    doc = fitz.open(caminho_pdf)
                    dialog.atualizar_progresso(30)

                    texto = "\n".join(page.get_text() for page in doc)
                    doc.close()
                    dialog.atualizar_progresso(50)

                    lancamentos = extrair_lancamentos(texto)
                    dialog.atualizar_progresso(70)

                    if not lancamentos:
                        QMessageBox.warning(
                            qt_parent, "Nenhum lançamento encontrado",
                            f"Nenhum lançamento válido encontrado em:\n{Path(caminho_pdf).name}"
                        )
                    else:
                        salvar_em_excel(caminho_pdf, lancamentos)
                        dialog.atualizar_progresso(90)

                except Exception as e:
                    QMessageBox.critical(
                        qt_parent, "Erro ao processar",
                        f"Erro ao processar o arquivo:\n{Path(caminho_pdf).name}\n\n{str(e)}"
                    )
                finally:
                    dialog.atualizar_progresso(100)
                    dialog.accept()

            continuar = qt_parent.mostrar_confirmacao(
                "Concluído",
                "Todos os arquivos selecionados foram processados.\n\nDeseja selecionar novos arquivos?"
            )
            if not continuar:
                break

    except Exception as erro_final:
        QMessageBox.critical(qt_parent, "Erro geral",
                             f"Erro inesperado:\n{str(erro_final)}")


# ==========================================================
# 🔹 Função de salvamento Excel (usada pelas duas versões)
# ==========================================================
def salvar_em_excel(caminho_pdf, lancamentos):
    # Remove a coluna 'Cor' já na criação do DataFrame
    df = pd.DataFrame(lancamentos, columns=[
                      "Data", "Histórico", "Valor", "Cor"])
    # mantém apenas as colunas desejadas
    df = df[["Data", "Histórico", "Valor"]]

    caminho_excel = Path(caminho_pdf).with_suffix('.xlsx')
    df.to_excel(caminho_excel, index=False)

    wb = load_workbook(caminho_excel)
    ws = wb.active

    # ======= ESTILO DO CABEÇALHO =======
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ======= APLICA COR AO VALOR =======
    for row in range(2, len(df) + 2):
        valor_cell = ws.cell(row=row, column=3)
        try:
            valor_texto = str(valor_cell.value).replace(
                "R$", "").replace(".", "").replace(",", ".").strip()
            valor_num = float(valor_texto)
            if valor_num > 0:
                valor_cell.font = Font(color="0000FF")  # azul
            elif valor_num < 0:
                valor_cell.font = Font(color="FF0000")  # vermelho
            valor_cell.alignment = Alignment(horizontal="right")
        except:
            continue

    # ======= AJUSTA LARGURA DAS COLUNAS =======
    for col in ws.columns:
        max_len = max(len(str(cell.value))
                      if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 2

    # ======= CRIA ESTILO DE TABELA =======
    num_linhas = ws.max_row
    tab = Table(displayName="TabelaLancamentos", ref=f"A1:C{num_linhas}")
    estilo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = estilo
    ws.add_table(tab)

    wb.save(caminho_excel)
