from main import LoaderDialog
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import pdfplumber
import re
import os


# ══════════════════════════════════════════════════════════════════════════════
# 🔹 Extração de lançamentos (mantida exatamente como está)
# ══════════════════════════════════════════════════════════════════════════════
def extrair_lancamentos(caminho_pdf):
    lancamentos = []
    padrao_data = re.compile(r"^(\d{2}/\d{2})")
    padrao_valor = re.compile(r"([\d\.]+,[\d]{2})(-?)")
    data_atual = ""
    capturando = False
    ano_extrato = None

    palavras_chave_inicio = [
        "data descrição entradas r$ saídas r$ saldo",
        "(créditos) (débitos)",
        "conta corrente | movimentação"
    ]

    palavras_chave_excluir = [
        "saldo", "saldo anterior", "realce", "extrato mensal", "notas explicativas",
        "limite contratado", "data da próxima renovação", "juros", "iof",
        "custo efetivo total", "totalizador de aplicações automáticas",
        "principal bruto líquido", "historico", "movimentação - aplicações",
        "resumo - mês", "na conta corrente", "principal", "total",
        "lis adicional", "lis recebíveis"
    ]

    with pdfplumber.open(caminho_pdf) as pdf:
        texto_completo = pdf.pages[0].extract_text() if pdf.pages else ""
        ano_match = re.search(r"\b(20\d{2})\b", texto_completo)
        ano_extrato = ano_match.group(
            1) if ano_match else str(datetime.now().year)

        for pagina in pdf.pages:
            palavras = pagina.extract_words(
                x_tolerance=1, y_tolerance=1, keep_blank_chars=False, use_text_flow=True
            )

            linhas_dict = {}
            for palavra in palavras:
                top = round(palavra["top"])
                if top not in linhas_dict:
                    linhas_dict[top] = []
                linhas_dict[top].append(palavra)

            for top in sorted(linhas_dict.keys()):
                palavras_linha = linhas_dict[top]
                linha_texto = " ".join(p["text"] for p in palavras_linha)
                linha_limpa = linha_texto.strip().lower()
                if not linha_limpa:
                    continue

                if not capturando:
                    if any(palavra in linha_limpa for palavra in palavras_chave_inicio):
                        capturando = True
                    continue

                if any(palavra in linha_limpa for palavra in palavras_chave_excluir):
                    continue

                if padrao_data.match(palavras_linha[0]["text"]):
                    data_curta = palavras_linha[0]["text"]
                    data_atual = f"{data_curta}/{ano_extrato}"
                    palavras_linha = palavras_linha[1:]

                valores = []
                descricao_tokens = []
                for p in palavras_linha:
                    if padrao_valor.match(p["text"]):
                        valores.append(p["text"])
                    else:
                        descricao_tokens.append(p["text"])

                descricao = " ".join(descricao_tokens)
                if not descricao.strip():
                    continue

                if valores and data_atual:
                    valor = valores[0].replace(".", "").replace(",", ".")
                    if valores[0].endswith("-"):
                        lancamentos.append(
                            [data_atual, descricao.strip(), -float(valor.rstrip("-"))])
                    else:
                        lancamentos.append(
                            [data_atual, descricao.strip(), float(valor)])

    return lancamentos


# ══════════════════════════════════════════════════════════════════════════════
# 🔹 Função de salvar em Excel (inalterada)
# ══════════════════════════════════════════════════════════════════════════════
def salvar_em_excel(dados, caminho_pdf):
    df = pd.DataFrame(dados, columns=["Data", "Descrição", "Valor"])
    nome_arquivo = os.path.splitext(os.path.basename(caminho_pdf))[0] + ".xlsx"
    pasta_destino = os.path.dirname(caminho_pdf)
    caminho_excel = os.path.join(pasta_destino, nome_arquivo)
    df.to_excel(caminho_excel, index=False)

    wb = load_workbook(caminho_excel)
    ws = wb.active

    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.col_idx == 3:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
                    cor = "0000FF" if cell.value > 0 else "FF0000"
                    cell.font = Font(color=cor)
                    cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    for col in ws.columns:
        max_length = max(len(str(cell.value))
                         if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    num_linhas = ws.max_row
    tab = Table(displayName="TabelaLancamentos", ref=f"A1:C{num_linhas}")

    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = estilo
    ws.add_table(tab)

    wb.save(caminho_excel)
    return caminho_excel


# ══════════════════════════════════════════════════════════════════════════════
# 🔹 Versão para Streamlit (Web)
# ══════════════════════════════════════════════════════════════════════════════
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    os.makedirs(output_dir, exist_ok=True)
    total = len(files)
    log_cb("Iniciando processamento dos extratos...")

    for i, caminho_pdf in enumerate(files, start=1):
        nome = os.path.basename(caminho_pdf)
        log_cb(f"📄 Processando {i}/{total}: {nome}")
        progress_cb(int((i - 1) / total * 70))

        try:
            dados = extrair_lancamentos(caminho_pdf)
            if not dados:
                log_cb(f"⚠️ Nenhum lançamento encontrado em {nome}")
                continue

            excel_path = os.path.join(
                output_dir, f"{os.path.splitext(nome)[0]}.xlsx")
            df = pd.DataFrame(dados, columns=["Data", "Descrição", "Valor"])
            df.to_excel(excel_path, index=False)
            salvar_em_excel(dados, excel_path)
            log_cb(
                f"✅ {len(dados)} lançamentos extraídos e salvos em {excel_path}")

        except Exception as e:
            log_cb(f"❌ Erro ao processar {nome}: {e}")

        progress_cb(int((i / total) * 100))

    progress_cb(100)
    log_cb("✅ Processamento concluído com sucesso! 🚀")


# ══════════════════════════════════════════════════════════════════════════════
# 🔹 Versão Desktop (inalterada)
# ══════════════════════════════════════════════════════════════════════════════
def processar_pdf_custom(janela):
    while True:
        arquivos, _ = QFileDialog.getOpenFileNames(
            janela, "Selecione um ou mais extratos PDF", "", "Arquivos PDF (*.pdf)"
        )
        if not arquivos:
            break

        for caminho_pdf in arquivos:
            dialog = LoaderDialog(janela, janela.light_theme)
            dialog.show()
            QApplication.processEvents()

            try:
                dialog.atualizar_progresso(10)
                dados = extrair_lancamentos(caminho_pdf)
                dialog.atualizar_progresso(60)

                if not dados:
                    QMessageBox.warning(
                        janela, "Aviso", f"Nenhum lançamento encontrado em:\n{caminho_pdf}"
                    )
                    dialog.atualizar_progresso(100)
                    dialog.accept()
                    continue

                salvar_em_excel(dados, caminho_pdf)
                dialog.atualizar_progresso(100)

            except Exception as e:
                QMessageBox.critical(
                    janela, "Erro", f"Erro ao processar {caminho_pdf}:\n{str(e)}"
                )
                dialog.atualizar_progresso(100)

            dialog.accept()

        continuar = janela.mostrar_confirmacao(
            "Concluído",
            "Todos os arquivos selecionados foram processados.\n\nDeseja selecionar novos arquivos?",
        )
        if not continuar:
            break
