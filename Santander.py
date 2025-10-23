# ==========================================================
# Módulo: Santander.py
# Compatível com:
#   - Streamlit (função processar_pdf_streamlit)
#   - PyQt5 (função processar_pdf_custom)
# ==========================================================

import os
import re
import time
import fitz
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from main import LoaderDialog


# ==========================================================
# 🔹 Extração de lançamentos do PDF Santander
# ==========================================================
def extrair_lancamentos_pdf(caminho_pdf):
    nome_base = os.path.splitext(os.path.basename(caminho_pdf))[0]
    caminho_debug = f"{nome_base}.debug.txt"

    doc = fitz.open(caminho_pdf)
    linhas_extraidas = []

    for pagina in doc:
        blocks = pagina.get_text("blocks")
        for b in blocks:
            for linha in b[4].split('\n'):
                linha_limpa = linha.strip()
                if linha_limpa:
                    linhas_extraidas.append(linha_limpa)

    # Gera arquivo de debug para análise
    with open(caminho_debug, "w", encoding="utf-8") as f:
        for linha in linhas_extraidas:
            f.write(linha + "\n")

    padrao_data = re.compile(r"^\d{2}/\d{2}/\d{4}$")
    padrao_valor = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}")
    padrao_documento = re.compile(r"\b\d{6}\b")

    blocos = []
    i = 0
    while i < len(linhas_extraidas):
        linha = linhas_extraidas[i]

        if padrao_data.match(linha):
            data = linha
            i += 1
            descricao_linhas = []
            valor = None
            documento = ""

            while i < len(linhas_extraidas):
                atual = linhas_extraidas[i].strip()

                if padrao_data.match(atual):
                    break

                if padrao_valor.search(atual):
                    valor_raw = padrao_valor.search(atual).group()
                    try:
                        valor = float(valor_raw.replace(
                            ".", "").replace(",", "."))
                    except ValueError:
                        valor = None

                    grupos6 = list(padrao_documento.finditer(atual))
                    if len(grupos6) >= 2:
                        documento = grupos6[1].group()
                        desc_fixa = atual[:grupos6[1].start()].strip()
                        descricao_linhas.append(desc_fixa)
                    else:
                        atual_sem_valor = atual.replace(valor_raw, "").strip()
                        descricao_linhas.append(atual_sem_valor)

                    i += 1
                    break
                else:
                    descricao_linhas.append(atual)
                    i += 1

            descricao_final = " ".join(descricao_linhas).strip()

            # Ignorar linhas de saldo
            if "saldo" in descricao_final.lower():
                continue

            if data and descricao_final and valor is not None:
                descricao_completa = f"{descricao_final} {documento}".strip()
                blocos.append((data, descricao_completa, valor))
        else:
            i += 1

    df = pd.DataFrame(blocos, columns=["Data", "Descrição", "Valor (R$)"])
    df["Data"] = pd.to_datetime(df["Data"], format="%d/%m/%Y", errors='coerce')
    df = df.dropna(subset=["Data"]).reset_index(drop=True)
    df["Data"] = df["Data"].dt.strftime("%d/%m/%Y")

    return df


# ==========================================================
# 🔹 Formatação e salvamento em Excel
# ==========================================================
def salvar_em_excel(df, caminho_excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lançamentos"

    bold = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")
    money_fmt = "#,##0.00"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.freeze_panes = 'A2'

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if r_idx == 1:
                cell.font = bold
                cell.alignment = center
                cell.fill = header_fill
            elif c_idx == 3:
                cell.number_format = money_fmt
                if isinstance(value, float):
                    if value < 0:
                        cell.font = Font(color="FF0000")
                    else:
                        cell.font = Font(color="0000FF")
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 16

    wb.save(caminho_excel)


# ==========================================================
# 💻 Modo Desktop (PyQt5)
# ==========================================================
def processar_pdf_custom(janela):
    while True:
        arquivos, _ = QFileDialog.getOpenFileNames(
            janela,
            "Selecione um ou mais extratos Santander (PDF)",
            "",
            "Arquivos PDF (*.pdf)"
        )

        if not arquivos:
            break

        for caminho_pdf in arquivos:
            dialog = LoaderDialog(janela, janela.light_theme)
            dialog.show()
            QApplication.processEvents()

            try:
                dialog.atualizar_progresso(10)
                df = extrair_lancamentos_pdf(caminho_pdf)
                dialog.atualizar_progresso(60)

                caminho_excel = Path(caminho_pdf).with_suffix('.xlsx')
                salvar_em_excel(df, caminho_excel)
                dialog.atualizar_progresso(100)

            except Exception as e:
                QMessageBox.critical(
                    janela, "Erro", f"Erro ao processar:\n{caminho_pdf}\n\n{str(e)}")

            dialog.accept()

        continuar = janela.mostrar_confirmacao(
            "Concluído",
            "Todos os arquivos foram processados.\n\nDeseja selecionar novos arquivos?"
        )
        if not continuar:
            break


# ==========================================================
# 🌐 Modo Web (Streamlit)
# ==========================================================
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    log_cb("Iniciando processamento de extratos Santander...")

    total = len(files)
    registros = []

    for i, pdf_path in enumerate(files, start=1):
        log_cb(f"Lendo arquivo {i}/{total}: {os.path.basename(pdf_path)}")
        time.sleep(0.3)
        progress_cb(int((i / total) * 70))

        try:
            df = extrair_lancamentos_pdf(pdf_path)
            if df.empty:
                log_cb(
                    f"⚠️ Nenhum lançamento encontrado em {os.path.basename(pdf_path)}")
                continue

            registros.extend(df.to_dict('records'))

        except Exception as e:
            log_cb(f"❌ Erro ao processar {os.path.basename(pdf_path)}: {e}")

    if registros:
        df_final = pd.DataFrame(registros)
        os.makedirs(output_dir, exist_ok=True)
        excel_path = os.path.join(output_dir, "Santander_Resultados.xlsx")
        df_final.to_excel(excel_path, index=False)
        log_cb(f"✅ Planilha gerada: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento encontrado nos PDFs enviados.")

    progress_cb(100)
    log_cb("Processamento concluído com sucesso! 🚀")
    return True
