# ==========================================================
# Módulo: Sicredi.py
# Compatível com:
#   - Streamlit (função processar_pdf_streamlit)
#   - PyQt5 (função processar_pdf_custom)
# ==========================================================

import os
import re
import time
import fitz  # PyMuPDF
import pandas as pd
from pathlib import Path
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from main import LoaderDialog


# ==========================================================
# 🔹 Extração dos lançamentos do PDF Sicredi
# ==========================================================
def extrair_lancamentos(pdf_path):
    doc = fitz.open(pdf_path)
    lancamentos = []

    padrao_data = re.compile(r"\d{2}/\d{2}/\d{4}")
    padrao_valor = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}$")

    buffer = {}
    for page in doc:
        linhas = page.get_text().split('\n')
        for linha in linhas:
            linha = linha.strip()

            if linha.upper() in ['DATA', 'DESCRIÇÃO', 'DOCUMENTO', 'VALOR (R$)', 'SALDO (R$)', 'SALDO ANTERIOR']:
                continue

            if padrao_valor.fullmatch(linha) and not buffer:
                continue

            if padrao_data.fullmatch(linha):
                buffer["DATA"] = linha
                continue

            if "DATA" in buffer and "DESCRIÇÃO" not in buffer:
                buffer["DESCRIÇÃO"] = linha
                continue

            if "DESCRIÇÃO" in buffer and "DOCUMENTO" not in buffer and not padrao_valor.fullmatch(linha):
                buffer["DOCUMENTO"] = linha
                continue

            if padrao_valor.fullmatch(linha):
                valor_str = linha.replace('.', '').replace(',', '.')
                try:
                    valor = float(valor_str)
                    buffer["VALOR"] = valor

                    if all(k in buffer for k in ["DATA", "DESCRIÇÃO", "VALOR"]):
                        lancamentos.append({
                            "DATA": buffer["DATA"],
                            "DESCRIÇÃO": buffer["DESCRIÇÃO"],
                            "DOCUMENTO": buffer.get("DOCUMENTO", ""),
                            "VALOR": buffer["VALOR"]
                        })
                except ValueError:
                    pass
                finally:
                    buffer = {}

    return lancamentos


# ==========================================================
# 🔹 Formatação e salvamento do Excel
# ==========================================================
def formatar_excel(caminho_excel):
    wb = load_workbook(caminho_excel)
    ws = wb.active

    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col == 4:
                cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                if isinstance(cell.value, (int, float)):
                    cell.font = Font(
                        color="0000FF" if cell.value >= 0 else "FF0000")
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    for col in range(1, ws.max_column + 1):
        max_length = max(
            len(str(ws.cell(row=row, column=col).value)) if ws.cell(
                row=row, column=col).value else 0
            for row in range(1, ws.max_row + 1)
        )
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = max_length + 2

    ref_final = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tab = Table(displayName="TabelaSicredi", ref=ref_final)
    estilo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    tab.tableStyleInfo = estilo
    ws.add_table(tab)

    wb.save(caminho_excel)
    return caminho_excel


# ==========================================================
# 💻 Modo Desktop (PyQt5)
# ==========================================================
def processar_pdf_custom(janela):
    while True:
        arquivos, _ = QFileDialog.getOpenFileNames(
            janela, "Selecione os arquivos PDF do Sicredi", "", "Arquivos PDF (*.pdf)"
        )
        if not arquivos:
            break

        for caminho_pdf in arquivos:
            dialog = LoaderDialog(janela, janela.light_theme)
            dialog.show()
            QApplication.processEvents()

            try:
                dialog.atualizar_progresso(10)

                dados = extrair_lancamentos(caminho_pdf)
                dialog.atualizar_progresso(40)

                if not dados:
                    QMessageBox.warning(
                        janela, "Atenção", f"Nenhum lançamento encontrado em:\n{caminho_pdf}")
                    dialog.atualizar_progresso(100)
                    dialog.accept()
                    continue

                df = pd.DataFrame(dados)
                caminho_excel = os.path.splitext(caminho_pdf)[0] + ".xlsx"
                df.to_excel(caminho_excel, index=False)
                dialog.atualizar_progresso(80)

                formatar_excel(caminho_excel)
                dialog.atualizar_progresso(100)

            except Exception as e:
                QMessageBox.critical(
                    janela, "Erro", f"Ocorreu um erro ao processar {caminho_pdf}:\n{e}")
                dialog.atualizar_progresso(100)

            dialog.accept()

        continuar = janela.mostrar_confirmacao(
            "Concluído",
            "Todos os arquivos selecionados foram processados.\n\nDeseja selecionar novos arquivos?"
        )
        if not continuar:
            break


# ==========================================================
# 🌐 Modo Web (Streamlit)
# ==========================================================
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    log_cb("Iniciando processamento dos extratos Sicredi...")

    total = len(files)
    registros = []

    for i, pdf_path in enumerate(files, start=1):
        log_cb(f"Lendo arquivo {i}/{total}: {os.path.basename(pdf_path)}")
        time.sleep(0.3)
        progress_cb(int((i / total) * 70))

        try:
            dados = extrair_lancamentos(pdf_path)
            if not dados:
                log_cb(
                    f"⚠️ Nenhum lançamento encontrado em {os.path.basename(pdf_path)}")
                continue

            registros.extend(dados)

        except Exception as e:
            log_cb(f"❌ Erro ao processar {os.path.basename(pdf_path)}: {e}")

    if registros:
        df_final = pd.DataFrame(registros)
        os.makedirs(output_dir, exist_ok=True)
        excel_path = os.path.join(output_dir, "Sicredi_Resultados.xlsx")
        df_final.to_excel(excel_path, index=False)
        log_cb(f"✅ Planilha gerada: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento encontrado nos PDFs enviados.")

    progress_cb(100)
    log_cb("Processamento concluído com sucesso! 🚀")
    return True
