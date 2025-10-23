# ==========================================================
# Módulo: XpInvestimentos.py
# Compatível com:
#   - Streamlit (função processar_pdf_streamlit)
#   - PyQt5 (função processar_pdf_custom)
# ==========================================================

import os
import re
import time
import pandas as pd
import fitz  # PyMuPDF
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QApplication
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ✅ Importa a classe LoaderDialog
from main import LoaderDialog


# ==========================================================
# 🔹 Extração de lançamentos do PDF
# ==========================================================
def extrair_lancamentos(pdf_path):
    doc = fitz.open(pdf_path)
    dados = []

    for page in doc:
        text = page.get_text("text")
        linhas = text.split('\n')

        buffer = ""
        data_atual = ""
        capturando = False

        for linha in linhas:
            linha = linha.strip()

            match_data = re.match(r"^(\d{2}/\d{2}/\d{4})\b", linha)
            if match_data:
                data_atual = match_data.group(1)
                buffer = linha[len(data_atual):].strip()
                capturando = True
                continue

            if capturando:
                buffer += " " + linha

            # Padrão 1: -R$ 1.234,56
            padrao1 = r"(.+?)\s+(-?)R\$\s*([\d\.]+,[\d]{2})\s+R\$"
            match1 = re.search(padrao1, buffer)

            if match1 and data_atual:
                historico = match1.group(1).strip()
                sinal = match1.group(2)
                valor = match1.group(3)
                valor_float = round(
                    float(valor.replace('.', '').replace(',', '.')), 2)
                if sinal == "-":
                    valor_float *= -1
                dados.append([data_atual, historico, valor_float])
                buffer = ""
                data_atual = ""
                capturando = False
                continue

            # Padrão 2: R$ -1.234,56
            padrao2 = r"(.+?)\s+R\$\s*(-?[\d\.]+,[\d]{2})\s+R\$"
            match2 = re.search(padrao2, buffer)

            if match2 and data_atual:
                historico = match2.group(1).strip()
                valor = match2.group(2)
                valor_float = round(
                    float(valor.replace('.', '').replace(',', '.')), 2)
                dados.append([data_atual, historico, valor_float])
                buffer = ""
                data_atual = ""
                capturando = False

    df = pd.DataFrame(dados, columns=["Data", "Histórico", "Valor (R$)"])
    return df


# ==========================================================
# 🔹 Formatação e salvamento Excel
# ==========================================================
def salvar_em_excel(df, pdf_path):
    excel_path = os.path.splitext(pdf_path)[0] + ".xlsx"
    df.to_excel(excel_path, index=False)

    wb = load_workbook(excel_path)
    ws = wb.active
    ws.freeze_panes = 'A2'

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Ajuste de colunas e formatação
    for col in ws.columns:
        max_length = max(len(str(cell.value))
                         if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_length + 2, 12)

    # Formatação da coluna de valor
    for row in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = thin_border
            if col_idx == 3:
                valor = cell.value
                if isinstance(valor, (int, float)):
                    cell.number_format = '#,##0.00'
                    cor = "0000FF" if valor >= 0 else "FF0000"
                    cell.font = Font(color=cor)
                    cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Criação da tabela formatada
    ref_final = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tab = Table(displayName="TabelaXpInvestimentos", ref=ref_final)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save(excel_path)
    return excel_path


# ==========================================================
# 💻 Modo Desktop (PyQt5)
# ==========================================================
def processar_pdf_custom(janela):
    while True:
        arquivos, _ = QFileDialog.getOpenFileNames(
            janela,
            "Selecione um ou mais PDFs da XP Investimentos",
            "",
            "Arquivos PDF (*.pdf)"
        )
        if not arquivos:
            break

        for pdf_path in arquivos:
            dialog = LoaderDialog(janela, janela.light_theme)
            dialog.show()
            QApplication.processEvents()

            try:
                dialog.atualizar_progresso(10)

                df = extrair_lancamentos(pdf_path)
                dialog.atualizar_progresso(50)

                if df.empty:
                    QMessageBox.warning(
                        janela, "Sem dados", f"Nenhum lançamento encontrado em:\n{pdf_path}")
                    dialog.atualizar_progresso(100)
                    dialog.accept()
                    continue

                salvar_em_excel(df, pdf_path)
                dialog.atualizar_progresso(100)

            except Exception as e:
                QMessageBox.critical(
                    janela, "Erro", f"Erro ao processar {pdf_path}:\n{str(e)}")
                dialog.atualizar_progresso(100)

            dialog.accept()

        continuar = janela.mostrar_confirmacao(
            "Concluído",
            "Todos os arquivos selecionados foram processados.\n\nDeseja selecionar novos arquivos?"
        )
        if not continuar:
            break


# ==========================================================
# 🌐 Modo Web (Streamlit)
# ==========================================================
def processar_pdf_streamlit(files, output_dir, progress_cb, log_cb):
    log_cb("Iniciando processamento dos extratos XP Investimentos...")

    total = len(files)
    registros = []

    for i, pdf_path in enumerate(files, start=1):
        log_cb(f"Lendo arquivo {i}/{total}: {os.path.basename(pdf_path)}")
        time.sleep(0.3)
        progress_cb(int((i / total) * 70))

        try:
            df = extrair_lancamentos(pdf_path)
            if df.empty:
                log_cb(
                    f"⚠️ Nenhum lançamento encontrado em {os.path.basename(pdf_path)}")
                continue

            registros.extend(df.values.tolist())

        except Exception as e:
            log_cb(f"❌ Erro ao processar {os.path.basename(pdf_path)}: {e}")

    if registros:
        df_final = pd.DataFrame(
            registros, columns=["Data", "Histórico", "Valor (R$)"])
        os.makedirs(output_dir, exist_ok=True)
        excel_path = os.path.join(
            output_dir, "XpInvestimentos_Resultados.xlsx")
        df_final.to_excel(excel_path, index=False)
        log_cb(f"✅ Planilha gerada: {excel_path}")
    else:
        log_cb("⚠️ Nenhum lançamento encontrado nos PDFs enviados.")

    progress_cb(100)
    log_cb("Processamento concluído com sucesso! 🚀")
    return True
